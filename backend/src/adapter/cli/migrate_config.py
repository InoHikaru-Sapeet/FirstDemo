"""xlsx → `config.json` 初期マイグレーション（T-14。設計書 §10）。

    make migrate-config                      # dry（既定）: 検証と差分レポートだけ
    make migrate-config ARGS="--apply"       # 検証を通ったら config.json を作成
    make migrate-config ARGS="--xlsx path/to/other.xlsx"

設計書 §10.3 の7手順をこの順で実行する:

1. xlsx の4シート（`情報カテゴリ`/`必須タグ`/`除外ルール`/`スコアリング軸`）を読む
2. 日本語 → ID（英小文字スネークケース）へ正規化（`中〜高` → `mid_high`。仕様書 §5.3）
3. `tunable_thresholds` に §5.2 の初期値を投入
4. モデル検証（T-04。JSON Schema と1:1）＋ クロスフィールド検証（T-05 の 1〜5）
5. §5.2 実データとの一致チェック（T-05 `validate_initial_config`＋件数 7/10/6/13）
6. `meta.revision=1` / `updated_by=null` / `updated_at=migration時刻` を打って出力
7. マイグレーションレポート（差分・警告）を出力

---

**動かしてはいけない点**

1. **dry が既定。** `--apply` を明示しない限り何も書かない。既存 `config.json` が
   あれば **revision を維持したまま差分だけ**報告する（設計書 §10.4 の冪等性）。
   既存の判断基準を黙って上書きしない。
2. **手順4-5 の検証に1件でも違反があれば書かずに中断する**（設計書 §10.4）。
   値を補正して通すことはしない（設計判断A）。
3. **書き込みは `ConfigRepository.create_initial()`（T-11）経由**。ここで直接
   `open()` しない。`revision=1` の採番・`config_revisions` への履歴記録・
   原子的書き込み（T-02）がすべてあちらに集約されているため、経路を増やすと
   「履歴に無い config.json」が生まれる。
4. **xlsx は初期投入元であって、以後の正ではない**（仕様書 §5）。このCLIは
   初回だけ動かすもので、2回目以降の変更は管理画面（`PUT /config`。T-13）が正。

---

**xlsx と仕様書 §5.2 の文言差分について（2026-08-14 決定。TASKS.md 要確認事項 #9）**

xlsx 実データと §5.2 は**件数・ID・配点・priority・severity・description が完全一致**
する一方、**7箇所だけ文言が違う**（§5.2 が xlsx を要約・略記している）。
**§5.2 を正とし、`SPEC_TEXT_NORMALIZATIONS` で寄せて、寄せたことを警告として
レポートに出す**方針で確定した。確定値（§5.2）は変更しない。
"""

import asyncio
from argparse import ArgumentParser, Namespace
from collections.abc import Callable
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path
from typing import Any, get_args

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from adapter.config_repository import (
    CONFIG_ADAPTER,
    INITIAL_REVISION,
    ConfigAlreadyExistsError,
    ConfigRepository,
    diff_configs,
)
from adapter.database.database import db_manager
from adapter.storage.artifact_store import CONFIG_FILENAME
from config import Settings, get_settings
from enterprise.entities.config import (
    EXCLUSION_RULE_COUNT,
    INFORMATION_CATEGORY_COUNT,
    REQUIRED_TAG_COUNT,
    SCHEMA_VERSION,
    SCORING_AXIS_COUNT,
    SCORING_TOTAL,
    AdoptionClass,
    CustomerRelevance,
    InfoType,
    PracticalUsability,
    Priority,
    Region,
    Reliability,
    Severity,
)
from enterprise.entities.config_validation import (
    ConfigIssue,
    ConfigIssueCode,
    validate_config,
    validate_initial_config,
)
from enterprise.entities.json_document import (
    DocumentIssue,
    DocumentParseError,
    validate_json_data,
)

# backend/src/adapter/cli/migrate_config.py → リポジトリルート
BACKEND_ROOT = Path(__file__).resolve().parents[3]
REPO_ROOT = BACKEND_ROOT.parent

# xlsx は「初期投入元＝確定値の出典」であって実行時入力ではないため、成果物
# （`artifact_root`）ではなく仕様書と同じ docs 配下に置く（2026-08-14 決定）。
SOURCE_XLSX_DIR = REPO_ROOT / "docs" / "source"
DEFAULT_XLSX = SOURCE_XLSX_DIR / "weekly_ai_intelligence_requirements.xlsx"

EXIT_OK = 0
# 既存 config があるので初期投入しなかった（§10.4 の冪等性。DB もファイルも無変更）。
EXIT_REFUSED = 1
# 手順4-5 の検証に失敗した（書き込んでいない）。
EXIT_VALIDATION_FAILED = 2
# xlsx が読めない・シートや列が足りない・未知の日本語表記。
EXIT_INVALID_INPUT = 3

SHEET_CATEGORIES = "情報カテゴリ"
SHEET_REQUIRED_TAGS = "必須タグ"
SHEET_EXCLUSION_RULES = "除外ルール"
SHEET_SCORING_AXES = "スコアリング軸"

# 見出し行は「No」で始まる行（1〜3行目はシートのタイトル・副題・空行）。行番号を
# 決め打ちにせず探すのは、xlsx 側にタイトル行が増減しても壊れないようにするため。
HEADER_KEY = "No"
# 「合計」行のラベル列（`No` が空なのでデータ行として読まれない）。
TOTAL_ROW_LABEL = "合計"

# 得点帯セルの区切り。`9-10:公式/政府一次情報` のように**帯の中に `/` を含む**ため
# 単独の `/` では割れない（区切りは前後に空白のある ` / `）。
BAND_SEPARATOR = " / "


class MigrationInputError(Exception):
    """xlsx 側の不備（シート・列の欠落、未知の日本語表記）。

    「読めなかった」と「検証に落ちた」を混ぜないために分けてある。前者は xlsx を
    直す話で、後者は §5.2 との突き合わせの話。
    """


# --- 日本語 → ID の対応表（手順2）--------------------------------------------
# ⚠️ **機械的な変換では作れないので表で持つ。** 「業務領域」→`business_area` の
# ように翻訳が必要で、かつ ID は中間xlsx の互換に直結する固定値（仕様書 §5.1）。
# 正は §5.2＝`enterprise.entities.config` の `Literal`。写し間違いはモデル検証
# （T-04）と ID 一致チェック（T-05 `check_fixed_identities`）で落ちる。

PRIORITY_BY_JA: dict[str, Priority] = {
    "高": Priority.HIGH,
    # 仕様書 §5.3。原本は波ダッシュ U+301C だが、Windows 経由で全角チルダ
    # U+FF5E に化けることがあるため両方を受ける（表記ゆれで落とさない）。
    "中〜高": Priority.MID_HIGH,
    "中～高": Priority.MID_HIGH,
    "中": Priority.MID,
    "低": Priority.LOW,
}

SEVERITY_BY_JA: dict[str, Severity] = {
    "完全除外": Severity.FULL_EXCLUDE,
    "原則除外": Severity.DEFAULT_EXCLUDE,
    "低優先": Severity.LOW_PRIORITY,
    "低優先または除外": Severity.LOW_PRIORITY_OR_EXCLUDE,
    "統合": Severity.MERGE,
}

# タグ名（xlsx）→ (`id`, `value_source`)。`value_source` は xlsx に列が無く、
# §5.2 が定めている値（参照先 enum が実在するかは T-05 が検証する）。
REQUIRED_TAG_BY_JA: dict[str, tuple[str, str]] = {
    "情報カテゴリ": ("information_category", "information_categories.id"),
    "AIテーマ": ("ai_theme", "free_controlled"),
    "業界": ("industry", "enums.industry"),
    "業務領域": ("business_area", "enums.business_area"),
    "情報種別": ("info_type", "enums.info_type"),
    "地域": ("region", "enums.region"),
    "信頼性": ("reliability", "enums.reliability"),
    "顧客関連度": ("customer_relevance", "enums.customer_relevance"),
    "実務活用可能性": ("practical_usability", "enums.practical_usability"),
    "レポート採用区分": ("adoption_class", "enums.adoption_class"),
}

# 軸名（xlsx）→ `id`。xlsx に ID 列が無いのはこのシートだけの事情ではなく、
# `情報カテゴリ` シートだけが ID 列を持っている（そちらは列の値をそのまま使う）。
SCORING_AXIS_BY_JA: dict[str, str] = {
    "顧客関連度": "customer_relevance",
    "実務活用可能性": "practical_usability",
    "AI業界・市場インパクト": "market_impact",
    "アドバイザリー活用度": "advisory_usability",
    "信頼性": "reliability",
    "緊急性・鮮度": "urgency_freshness",
}

# --- xlsx と §5.2 の文言差分（2026-08-14 決定：§5.2 を正・警告として報告）----
# ⚠️ **ここに「意味を変える」変換を足さないこと。** 対象は §5.2 が xlsx を要約・
# 略記した7箇所だけで、確定値（§5.2）へ寄せるための表。適用したことは必ず警告に
# 出す（黙って書き換えると、xlsx を見た人が config.json を読めなくなる）。
# xlsx 側が §5.2 と揃うように直された場合、対応する行は発火しなくなる。
# **使われなかった行はレポートで「未適用」として報告する**（腐った行を残さない）。
SPEC_TEXT_NORMALIZATIONS: dict[str, str] = {
    # required_tags[].purpose（No.1 / 2 / 4 / 6 / 9 / 10）— §5.2 は語尾を落としている
    "レポート全体の分類軸になる": "レポート全体の分類軸",
    "検索・絞り込みの中心になる": "検索・絞り込みの中心",
    "顧客の実務テーマと接続する": "顧客の実務テーマと接続",
    "日本/海外/グローバルの区別に必要": "日本/海外/グローバルの区別",
    "ニュースを示唆に変えるために必要": "ニュースを示唆に変える",
    "3出力形式への振り分けに必要": "3出力形式への振り分け",
    # scoring_axes[reliability].bands[2] — §5.2 は「プレスリリース」を「PR」と略記
    "5-6:ブログ・プレスリリース要確認": "5-6:ブログ・PR要確認",
}

# --- xlsx に無く §5.2 から投入するブロック（手順3）--------------------------
# `enums` のうち型で固定できるものは `Literal` から起こす（正を1箇所に保つ）。
# `industry` / `business_area` は運用で増減しうる自由文字列（T-04）なので型から
# 起こせず、`source_whitelist_hint` と `tunable_thresholds` も xlsx に無い。
# これらは §5.2 の逐語コピーで、**一致は T-05 の `validate_initial_config`
# （tunable）とテスト（§5.2 の実データとの全体比較）が独立に確かめる。**

SPEC_INDUSTRIES: tuple[str, ...] = (
    "業界横断",
    "不動産",
    "製造",
    "モビリティ・自動車",
    "情報通信",
    "IT",
    "半導体",
    "金融",
    "小売",
    "物流",
    "エネルギー",
    "メディア・エンタメ",
    "公共",
    "教育",
    "医薬品",
    "通信",
    "ロボティクス",
    "クラウド",
    "航空宇宙",
)

SPEC_BUSINESS_AREAS: tuple[str, ...] = (
    "AI戦略",
    "ガバナンス",
    "法務・コンプライアンス",
    "情報システム",
    "セキュリティ",
    "開発",
    "業務プロセス改革",
    "マーケティング",
    "営業",
    "カスタマーサポート",
    "バックオフィス",
    "人材育成・組織変革",
    "研究開発",
    "調達",
    "データ基盤",
    "生産・現場オペレーション",
    "コンテンツ制作",
    "経営企画",
)

SPEC_SOURCE_WHITELIST_HINT: tuple[str, ...] = (
    "TechCrunch",
    "VentureBeat",
    "Ledge.ai",
    "ITmedia",
    "公式プレスリリース",
    "政府・公的機関",
)

SPEC_TUNABLE_THRESHOLDS: dict[str, Any] = {
    "min_total_score_to_publish": 60,
    "adoption_class_score_map": {
        "propose_next_meeting": 85,
        "reference_info": 70,
        "share_only": 60,
    },
    "min_reliability_score_to_publish": 5,
    "weekly": {
        "target_industry": "不動産",
        "max_industry_topics": 5,
        "max_common_topics": 6,
        "point_of_week_required": True,
    },
    "monthly": {
        "target_case_count": 15,
        "chapter_count_hint": 5,
        "min_score_for_case": 80,
        "require_editorial_and_closing": True,
    },
    "dedup": {
        "lookback_weeks": 8,
        "title_similarity_threshold": 0.85,
        "treat_same_url_as_duplicate": True,
        # ⚠️ **仕様書 §5.2 に無い鍵**（2026-08-16 の決定2。§11.1 の月次
        # 「直近数ヶ月」＝3。→ TASKS.md T-21 備考・T-38）。
        "monthly_lookback_months": 3,
    },
}

# §5.2 の `meta`。`revision` / `updated_at` / `updated_by` は
# `ConfigRepository.create_initial()` が打つ（T-11）ので、ここでは初期値を置くだけ。
CONFIG_NAME = "ai_intelligence_requirements"
SOURCE_OF_TRUTH_XLSX = "weekly_ai_intelligence_requirements.xlsx"
ADMIN_ONLY: tuple[str, ...] = ("admin",)

# 期待件数（仕様書 §5.1 の確定値）。モデル（T-04）が min/max_length で担保して
# いるが、**レポートに件数を出すこと自体が手順5の受け入れ材料**なので明示で持つ。
EXPECTED_COUNTS: dict[str, int] = {
    "information_categories": INFORMATION_CATEGORY_COUNT,
    "required_tags": REQUIRED_TAG_COUNT,
    "scoring_axes": SCORING_AXIS_COUNT,
    "exclusion_rules": EXCLUSION_RULE_COUNT,
}


# --- レポート（手順7）--------------------------------------------------------


@dataclass(frozen=True)
class MigrationWarning:
    """「そのままでは §5.2 と違うので寄せた」等、書き込みは止めない気づき。

    Attributes:
        path: config ルートからのドット区切りパス（`ConfigIssue.path` と同じ表記）
        reason: 何をどう扱ったか
    """

    path: str
    reason: str


@dataclass
class MigrationReport:
    """マイグレーション1回の結果（設計書 §10.3 手順7）。"""

    xlsx: Path
    config_path: Path
    apply: bool
    counts: dict[str, int] = field(default_factory=dict)
    warnings: list[MigrationWarning] = field(default_factory=list)
    parse_issues: list[DocumentIssue] = field(default_factory=list)
    issues: list[ConfigIssue] = field(default_factory=list)
    # 既存 config があるときの差分（`meta.*` は除く。T-11 の `diff_configs`）
    diff: dict[str, dict[str, Any]] | None = None
    existing_revision: int | None = None
    written_revision: int | None = None

    @property
    def ok(self) -> bool:
        return not self.parse_issues and not self.issues


def render_report(report: MigrationReport) -> str:
    """レポートを人が読む形に整える。CLI の標準出力に出す唯一の成果物。"""
    mode = "apply（書き込み）" if report.apply else "dry（既定・書き込みなし）"
    lines = [
        "=== config.json 初期マイグレーション（T-14 / 設計書 §10）===",
        f"モード      : {mode}",
        f"入力 xlsx   : {report.xlsx}",
        f"出力 config : {report.config_path}",
    ]

    if report.counts:
        counts = " / ".join(
            f"{name}={count}（期待{EXPECTED_COUNTS[name]}）"
            for name, count in report.counts.items()
        )
        lines += ["", f"件数        : {counts}"]

    if report.parse_issues:
        lines += ["", f"■ 構造検証エラー {len(report.parse_issues)}件（手順4。T-04）"]
        lines += [f"  - {i.path}: {i.reason}" for i in report.parse_issues]

    if report.issues:
        lines += ["", f"■ 検証エラー {len(report.issues)}件（手順4-5。T-05）"]
        lines += [f"  - [{i.code}] {i.path}: {i.reason}" for i in report.issues]

    if report.warnings:
        lines += ["", f"■ 警告 {len(report.warnings)}件"]
        lines += [f"  - {w.path}: {w.reason}" for w in report.warnings]

    if report.existing_revision is not None:
        lines += [
            "",
            f"■ 既存 config あり（revision={report.existing_revision} を維持）",
        ]
        if report.diff:
            lines.append(f"  差分 {len(report.diff)}件（meta.* を除く）:")
            lines += [
                f"    - {path}: {change['before']!r} → {change['after']!r}"
                for path, change in report.diff.items()
            ]
        else:
            lines.append("  差分なし（xlsx から起こした内容と一致）")

    lines += [""]
    if report.written_revision is not None:
        lines.append(
            f"→ 作成しました: {report.config_path}（revision={report.written_revision}"
            " / updated_by=null）"
        )
    elif not report.ok:
        lines.append("→ 検証に失敗したため書き込んでいません（設計書 §10.4）。")
    elif report.existing_revision is not None:
        lines.append(
            "→ 既存 config があるため書き込んでいません。"
            "変更は管理画面（PUT /config）から行ってください。"
        )
    else:
        lines.append(
            "→ dry モードなので書き込んでいません。"
            "作成するには --apply を付けて実行してください。"
        )
    return "\n".join(lines)


# --- 手順1-2: xlsx を読んで正規化 --------------------------------------------


def _cell(value: object) -> str:
    """セルを文字列にする。空セルは空文字。前後の空白は落とす。"""
    return "" if value is None else str(value).strip()


def _sheet(workbook: Any, name: str) -> Worksheet:
    if name not in workbook.sheetnames:
        raise MigrationInputError(
            f"シート {name!r} がありません（必要なのは: "
            f"{SHEET_CATEGORIES} / {SHEET_REQUIRED_TAGS} / "
            f"{SHEET_EXCLUSION_RULES} / {SHEET_SCORING_AXES}）"
        )
    return workbook[name]


def _read_rows(sheet: Worksheet, *, columns: tuple[str, ...]) -> list[dict[str, str]]:
    """見出し行を探し、指定列だけを取り出す。

    列の**順序に依存しない**（見出し名で引く）。`No` が空の行は読まない＝
    `スコアリング軸` シート末尾の「合計」行は自然に除かれる。

    Args:
        sheet: 対象シート
        columns: 必要な見出し名

    Returns:
        1行 = 見出し名 → 値 の辞書。`No` も含む

    Raises:
        MigrationInputError: 見出し行が無い / 必要な列が足りない
    """
    rows = list(sheet.iter_rows(values_only=True))
    header_index = next(
        (i for i, row in enumerate(rows) if row and _cell(row[0]) == HEADER_KEY),
        None,
    )
    if header_index is None:
        raise MigrationInputError(
            f"シート {sheet.title!r} に見出し行（1列目が {HEADER_KEY!r}）がありません"
        )

    header = [_cell(value) for value in rows[header_index]]
    missing = [name for name in columns if name not in header]
    if missing:
        raise MigrationInputError(
            f"シート {sheet.title!r} に列 {', '.join(missing)} がありません"
            f"（見出し: {', '.join(name for name in header if name)}）"
        )

    index_of = {name: header.index(name) for name in (HEADER_KEY, *columns)}
    return [
        {name: _cell(row[position]) for name, position in index_of.items()}
        for row in rows[header_index + 1 :]
        if row and _cell(row[0])
    ]


def _lookup[T](table: dict[str, T], value: str, *, sheet: str, column: str) -> T:
    """日本語表記を ID へ引く。未知の表記は**推測せず**落とす。"""
    if value not in table:
        raise MigrationInputError(
            f"{sheet} シートの {column} に未知の表記 {value!r} があります"
            f"（対応表にあるのは: {' / '.join(table)}）"
        )
    return table[value]


def _int(value: str, *, sheet: str, column: str) -> int:
    try:
        return int(float(value))
    except ValueError as exc:
        raise MigrationInputError(
            f"{sheet} シートの {column} が数値ではありません: {value!r}"
        ) from exc


class _Normalizer:
    """§5.2 へ寄せる文言正規化と、その記録（`SPEC_TEXT_NORMALIZATIONS`）。"""

    def __init__(self) -> None:
        self.warnings: list[MigrationWarning] = []
        self._applied: set[str] = set()

    def warn(self, path: str, reason: str) -> None:
        self.warnings.append(MigrationWarning(path=path, reason=reason))

    def text(self, value: str, *, path: str) -> str:
        replacement = SPEC_TEXT_NORMALIZATIONS.get(value)
        if replacement is None:
            return value
        self._applied.add(value)
        self.warn(
            path,
            f"xlsx の文言 {value!r} を仕様書 §5.2 の確定値 {replacement!r} へ"
            "寄せた（§5.2 を正とする決定。TASKS.md 要確認事項 #9）",
        )
        return replacement

    def report_unused(self) -> None:
        """一度も使われなかった正規化行を警告に出す（腐った行を残さない）。"""
        for value in SPEC_TEXT_NORMALIZATIONS:
            if value in self._applied:
                continue
            self.warn(
                "SPEC_TEXT_NORMALIZATIONS",
                f"正規化行 {value!r} が xlsx に見つからなかった（未適用）。"
                "xlsx が §5.2 に合わせて修正されたなら、この行は削除できる",
            )


def _build_information_categories(
    sheet: Worksheet, normalizer: _Normalizer
) -> list[dict[str, Any]]:
    """`情報カテゴリ` → `information_categories[]`（設計書 §10.2）。

    このシートだけ `ID` 列を持つので、ID は表に頼らず列の値を使う（値が §5.2 の
    確定 ID と揃っているかはモデル検証と T-05 の ID 一致チェックが見る）。
    """
    columns = ("カテゴリ名", "ID", "説明", "初期優先度")
    return [
        {
            "id": row["ID"],
            "label": normalizer.text(
                row["カテゴリ名"], path=f"information_categories.{index}.label"
            ),
            "priority": _lookup(
                PRIORITY_BY_JA,
                row["初期優先度"],
                sheet=SHEET_CATEGORIES,
                column="初期優先度",
            ).value,
            "description": normalizer.text(
                row["説明"], path=f"information_categories.{index}.description"
            ),
        }
        for index, row in enumerate(_read_rows(sheet, columns=columns))
    ]


def _build_required_tags(
    sheet: Worksheet, normalizer: _Normalizer
) -> list[dict[str, Any]]:
    """`必須タグ` → `required_tags[]`（設計書 §10.2）。

    `id` / `value_source` は xlsx に列が無いので `REQUIRED_TAG_BY_JA` から引く。
    `required` は全10件 true（仕様書 §5.1「10タグを漏れなく保持」・型は
    `Literal[True]`）。
    """
    columns = ("タグ名", "型", "説明・目的")
    tags: list[dict[str, Any]] = []
    for index, row in enumerate(_read_rows(sheet, columns=columns)):
        tag_id, value_source = _lookup(
            REQUIRED_TAG_BY_JA,
            row["タグ名"],
            sheet=SHEET_REQUIRED_TAGS,
            column="タグ名",
        )
        tags.append(
            {
                "id": tag_id,
                "label": row["タグ名"],
                "type": row["型"],
                "required": True,
                "purpose": normalizer.text(
                    row["説明・目的"], path=f"required_tags.{index}.purpose"
                ),
                "value_source": value_source,
            }
        )
    return tags


def _build_scoring_axes(
    sheet: Worksheet, normalizer: _Normalizer
) -> tuple[list[dict[str, Any]], int | None]:
    """`スコアリング軸` → `scoring_axes[]`＋`scoring_total`（設計書 §10.2）。

    Returns:
        (軸のリスト, xlsx の「合計」行の値。読めなければ None)
    """
    columns = ("軸名", "配点", "評価観点", "得点帯と基準")
    axes: list[dict[str, Any]] = []
    for index, row in enumerate(_read_rows(sheet, columns=columns)):
        axes.append(
            {
                "id": _lookup(
                    SCORING_AXIS_BY_JA,
                    row["軸名"],
                    sheet=SHEET_SCORING_AXES,
                    column="軸名",
                ),
                "label": row["軸名"],
                "weight": _int(row["配点"], sheet=SHEET_SCORING_AXES, column="配点"),
                "criterion": normalizer.text(
                    row["評価観点"], path=f"scoring_axes.{index}.criterion"
                ),
                "bands": [
                    normalizer.text(band, path=f"scoring_axes.{index}.bands")
                    for band in row["得点帯と基準"].split(BAND_SEPARATOR)
                    if band.strip()
                ],
            }
        )
    return axes, _read_total_row(sheet, normalizer)


def _read_total_row(sheet: Worksheet, normalizer: "_Normalizer") -> int | None:
    """「合計」行の配点を読む（xlsx 自身が主張する満点）。

    読めなければ **警告に留めて None を返す**（中断しない）。原本の合計セルは
    `=SUM(C5:C10)` の数式で、数式の計算結果が保存されていない xlsx（openpyxl で
    書き出したもの等）では値が取れない。**満点の主張が読めないことと、満点が
    違うことは別**で、後者だけを検証で落としたい（`_check_declared_total`）。

    Returns:
        合計値。読めなければ None
    """
    for row in sheet.iter_rows(values_only=True):
        cells = [_cell(value) for value in row]
        if TOTAL_ROW_LABEL not in cells:
            continue
        position = cells.index(TOTAL_ROW_LABEL) + 1
        value = cells[position] if position < len(cells) else ""
        try:
            return _int(value, sheet=SHEET_SCORING_AXES, column=TOTAL_ROW_LABEL)
        except MigrationInputError:
            normalizer.warn(
                "scoring_total",
                f"「{TOTAL_ROW_LABEL}」行の値 {value!r} が数値として読めないため、"
                "満点の突き合わせを省略した（数式の計算結果が保存されていない xlsx）。"
                f"軸 weight の合計が {SCORING_TOTAL} であることは T-05 が検証する",
            )
            return None

    normalizer.warn(
        "scoring_total",
        f"シート {SHEET_SCORING_AXES!r} に「{TOTAL_ROW_LABEL}」行が無いため、"
        "満点の突き合わせを省略した",
    )
    return None


def _build_exclusion_rules(
    sheet: Worksheet, normalizer: _Normalizer
) -> list[dict[str, Any]]:
    """`除外ルール` → `exclusion_rules[]`（設計書 §10.2）。

    `enabled` は全13件 true（§5.2 の初期値。T-05 が一致を確かめる）。
    """
    columns = ("除外区分", "ルール名", "具体例")
    return [
        {
            "no": _int(row[HEADER_KEY], sheet=SHEET_EXCLUSION_RULES, column=HEADER_KEY),
            "severity": _lookup(
                SEVERITY_BY_JA,
                row["除外区分"],
                sheet=SHEET_EXCLUSION_RULES,
                column="除外区分",
            ).value,
            "enabled": True,
            "name": normalizer.text(
                row["ルール名"], path=f"exclusion_rules.{index}.name"
            ),
            "examples": normalizer.text(
                row["具体例"], path=f"exclusion_rules.{index}.examples"
            ),
        }
        for index, row in enumerate(_read_rows(sheet, columns=columns))
    ]


def _spec_enums() -> dict[str, list[str]]:
    """`enums`（xlsx に無い。§5.2）。

    型（T-04）で固定されているものは `Literal` / `StrEnum` から起こす。写しを
    2つ持たないため。`industry` / `business_area` は自由文字列なので §5.2 の
    逐語コピー。
    """
    return {
        "priority": [priority.value for priority in Priority],
        "severity": [severity.value for severity in Severity],
        "reliability": list(get_args(Reliability)),
        "customer_relevance": list(get_args(CustomerRelevance)),
        "practical_usability": list(get_args(PracticalUsability)),
        "adoption_class": list(get_args(AdoptionClass)),
        "region": list(get_args(Region)),
        "info_type": list(get_args(InfoType)),
        "industry": list(SPEC_INDUSTRIES),
        "business_area": list(SPEC_BUSINESS_AREAS),
    }


def build_config_payload(
    xlsx: Path, *, revision: int, updated_at: datetime
) -> tuple[dict[str, Any], list[MigrationWarning], int | None]:
    """xlsx から `config.json` 相当の生データを組み立てる（手順1〜3・6）。

    キー順は §5.2 と同じ（`IntelligenceConfig` の宣言順）。revision 間の diff が
    読める形を保つため。

    Args:
        xlsx: 入力ファイル
        revision: `meta.revision`（初期投入は 1。dry で既存があればその値）
        updated_at: `meta.updated_at`（migration 時刻）

    Returns:
        (生データ, 警告, xlsx の合計行の値。読めなければ None)

    Raises:
        MigrationInputError: 読み込み・正規化に失敗した場合
    """
    if not xlsx.is_file():
        raise MigrationInputError(
            f"xlsx がありません: {xlsx}\n"
            "元ファイルの置き場は docs/source/ です（--xlsx で変更できます）。"
        )

    try:
        workbook = load_workbook(xlsx, data_only=True)
    except Exception as exc:  # openpyxl は壊れたファイルで多様な例外を投げる
        raise MigrationInputError(f"xlsx を開けません: {xlsx}（{exc}）") from exc

    try:
        normalizer = _Normalizer()
        categories = _build_information_categories(
            _sheet(workbook, SHEET_CATEGORIES), normalizer
        )
        tags = _build_required_tags(_sheet(workbook, SHEET_REQUIRED_TAGS), normalizer)
        axes, declared_total = _build_scoring_axes(
            _sheet(workbook, SHEET_SCORING_AXES), normalizer
        )
        rules = _build_exclusion_rules(
            _sheet(workbook, SHEET_EXCLUSION_RULES), normalizer
        )
    finally:
        workbook.close()

    normalizer.report_unused()

    payload = {
        "schema_version": SCHEMA_VERSION,
        "meta": {
            "config_name": CONFIG_NAME,
            "source_of_truth_xlsx": SOURCE_OF_TRUTH_XLSX,
            "editable_by": list(ADMIN_ONLY),
            "visible_to": list(ADMIN_ONLY),
            "updated_at": updated_at.isoformat(),
            "updated_by": None,
            "revision": revision,
        },
        "information_categories": categories,
        "required_tags": tags,
        "scoring_axes": axes,
        "scoring_total": SCORING_TOTAL,
        "exclusion_rules": rules,
        "enums": _spec_enums(),
        "tunable_thresholds": SPEC_TUNABLE_THRESHOLDS,
        "source_whitelist_hint": list(SPEC_SOURCE_WHITELIST_HINT),
    }
    return payload, normalizer.warnings, declared_total


# --- 手順4-5: 検証 -----------------------------------------------------------


def _check_declared_total(
    declared_total: int | None, axes_total: int
) -> list[ConfigIssue]:
    """xlsx の「合計」行が §5.2 の満点（100）と一致するか。

    `scoring_total` は `Literal[100]` で固定なので、xlsx 側が別の満点を主張して
    いたら**そのまま 100 として通してはいけない**（配点の意味が変わる）。
    軸の weight 合計との一致は T-05 `check_weight_sum` が見る。
    None（合計行が読めなかった）は警告済みなので何も言わない。
    """
    if declared_total is None or declared_total == SCORING_TOTAL:
        return []
    return [
        ConfigIssue(
            path="scoring_total",
            reason=(
                f"xlsx の「{TOTAL_ROW_LABEL}」行が {declared_total} で、"
                f"仕様書 §5.2 の満点 {SCORING_TOTAL} と一致しない"
                f"（軸の weight 合計は {axes_total}）"
            ),
            code=ConfigIssueCode.INITIAL_VALUE_MISMATCH,
        )
    ]


# --- 実行 --------------------------------------------------------------------


async def run(
    repo: ConfigRepository,
    *,
    xlsx: Path = DEFAULT_XLSX,
    apply: bool = False,
    updated_at: datetime | None = None,
    out: Callable[[str], None] = print,
) -> int:
    """CLI の本体。終了コードを返す（例外で落ちない）。

    Args:
        repo: config の読み書き口（T-11）。**直接 open() はしない**
        xlsx: 入力 xlsx
        apply: True なら検証を通ったときだけ書き込む。False（既定）は dry
        updated_at: `meta.updated_at`。既定は現在時刻（migration 時刻）
        out: 出力先（テストで差し替える）

    Returns:
        `EXIT_OK` / `EXIT_REFUSED` / `EXIT_VALIDATION_FAILED` / `EXIT_INVALID_INPUT`
    """
    existing = repo.load() if repo.exists() else None
    report = MigrationReport(
        xlsx=xlsx,
        config_path=repo.path,
        apply=apply,
        existing_revision=existing.meta.revision if existing else None,
    )

    # 既存があれば revision を維持する（設計書 §10.4。差分レポートで revision が
    # 1 へ戻ったように見えるのを防ぐ）。
    revision = existing.meta.revision if existing else INITIAL_REVISION

    try:
        payload, warnings, declared_total = build_config_payload(
            xlsx,
            revision=revision,
            updated_at=updated_at or datetime.now(tz=get_settings().tzinfo),
        )
    except MigrationInputError as exc:
        out(f"中止しました: {exc}")
        return EXIT_INVALID_INPUT

    report.warnings = warnings
    report.counts = {name: len(payload[name]) for name in EXPECTED_COUNTS}

    # 手順4: モデル検証（T-04。JSON Schema と1:1。件数 7/10/6/13・ID・値域はここ）
    try:
        candidate = validate_json_data(CONFIG_ADAPTER, payload, label=CONFIG_FILENAME)
    except DocumentParseError as exc:
        report.parse_issues = exc.issues
        out(render_report(report))
        return EXIT_VALIDATION_FAILED

    # 手順4: クロスフィールド検証（T-05 の 1〜5）／ 手順5: §5.2 実データとの一致
    report.issues = [
        *validate_config(candidate),
        *validate_initial_config(candidate),
        *_check_declared_total(
            declared_total, sum(axis.weight for axis in candidate.scoring_axes)
        ),
    ]
    if not report.ok:
        out(render_report(report))
        return EXIT_VALIDATION_FAILED

    if existing is not None:
        report.diff = diff_configs(existing, candidate)
        out(render_report(report))
        # 既存があるのに --apply された場合だけ拒否として扱う（dry は正常終了）。
        return EXIT_REFUSED if apply else EXIT_OK

    if not apply:
        out(render_report(report))
        return EXIT_OK

    # 手順6: 書き込みは T-11 経由（revision=1 / updated_by=null / updated_at を打つ）
    try:
        written = await repo.create_initial(candidate)
    except ConfigAlreadyExistsError as exc:  # 直前に別経路が作った場合
        out(f"中止しました: {exc}")
        return EXIT_REFUSED

    report.written_revision = written.meta.revision
    out(render_report(report))
    return EXIT_OK


def _build_parser() -> ArgumentParser:
    parser = ArgumentParser(
        prog="migrate-config",
        description=(
            "xlsx から config.json を初期投入する（T-14 / 設計書 §10）。"
            "既定は dry で、--apply を付けたときだけ書き込む。"
        ),
    )
    parser.add_argument(
        "--xlsx",
        type=Path,
        default=DEFAULT_XLSX,
        help=f"入力 xlsx（既定: {DEFAULT_XLSX}）",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="検証を通ったら config.json を作成する（既定は dry で書き込まない）",
    )
    return parser


def main(argv: list[str] | None = None) -> int:
    args: Namespace = _build_parser().parse_args(argv)

    async def _run() -> int:
        # dry でも DB セッションを開く（`ConfigRepository` が改訂履歴も扱うため）。
        # ただし dry が触るのは `config.json` の読み込みだけで、DB への書き込みは
        # `--apply` のとき（`create_initial()`）にしか起きない。
        settings: Settings = get_settings()
        async with db_manager.session() as db:
            repo = ConfigRepository.from_settings(db, settings)
            return await run(repo, xlsx=args.xlsx, apply=args.apply)

    try:
        return asyncio.run(_run())
    except KeyboardInterrupt:
        print("中止しました: 中断されました。")
        return EXIT_INVALID_INPUT


if __name__ == "__main__":
    raise SystemExit(main())


__all__ = [
    "DEFAULT_XLSX",
    "EXIT_INVALID_INPUT",
    "EXIT_OK",
    "EXIT_REFUSED",
    "EXIT_VALIDATION_FAILED",
    "MigrationInputError",
    "MigrationReport",
    "MigrationWarning",
    "build_config_payload",
    "main",
    "render_report",
    "run",
]
