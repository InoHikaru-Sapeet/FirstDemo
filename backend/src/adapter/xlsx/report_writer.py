"""中間xlsx のライタ／リーダ（設計書 §2.2 ／ 仕様書 §8 ／ 設計判断B ／ T-22）。

中間xlsx は「ファイルが正」方針の受け渡し単位で、**HTML生成（T-24 / T-25）の入力**
であり、**重複判定（§11.1）の参照元**でもある。この モジュールがその唯一の
読み書き口。

| ファイル | シート | 列 |
|---|---|---|
| `weekly_ai_intelligence_report.xlsx` | ISO週ごとに1枚 ＋ `除外ログ` | 22列 / 6列 |
| `monthly_ai_leading_cases.xlsx` | 対象月ごとに1枚 | 8列 |

---

**⚠️ 列名・列順はここに書かない。** すべて T-07（`enterprise.entities.report_columns`）
の定義から引く。セル値の書き出し・読み戻し（日付の表記・multi の `;` 区切り・
月次「解説」の `\\n\\n`）も T-07 の `format_row()` / `parse_row()` に任せる。
ライタとリーダが別々に列順を持つと、片方を直したときにもう片方が黙って壊れる。

---

**設計判断B（正規名は上書き ＋ 履歴退避）**

正規名（固定ファイル名・固定シート名）は upsert する。PROMPT-3 相当の後段が
固定名を入力に読むため（§13.4）、名前を変えていく方式は採れない。そのかわり
**上書きの直前に、その時点のファイルを `_history/{period}/{revision}_{run_id}/` へ
退避する**（T-02 の `ArtifactStore.archive()`。世代上限つき）。

**同じ period の再実行は、その period のシートを丸ごと作り直す**（追記しない）。
追記だと2回目の実行で記事が二重に並ぶ（§14 冪等性）。

⚠️ **`除外ログ` シートだけは append**（§8.1 / T-22 完了条件）。**週次ブックの1枚に
一本化**し、月次実行で出た除外もここへ積む（T-21 備考。参照元を2箇所に割らない）。
その結果、同じ period を2回実行すると除外ログには同じ行が2回並ぶ——
**「いつ何を落としたか」の記録**なので消さない（本編の重複とは意味が違う）。
"""

import logging
from collections.abc import Iterable, Mapping, Sequence
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet

from adapter.storage.artifact_store import ArtifactStore
from enterprise.entities.period import (
    Period,
    PeriodError,
    monthly_period_of,
    parse_period,
    weekly_period_of,
)
from enterprise.entities.report_columns import (
    EXCLUSION_LOG_SHEET,
    EXCLUSION_LOG_SHEET_NAME,
    MONTHLY_CASE_SHEET,
    WEEKLY_ARTICLE_SHEET,
    ReportColumn,
    SheetLayout,
    format_row,
    header_row,
    parse_row,
)
from enterprise.services.dedup import DedupHistory, KnownArticle, KnownOrigin

logger = logging.getLogger(__name__)

# 週次の各週シートの前置き（仕様書 §8.1。**逐語**）。
# 1行目タイトル / 2行目説明 / 3行目空行 / 4行目ヘッダ / 5行目以降データ。
WEEKLY_SHEET_TITLE_FORMAT = "Weekly AI Intelligence レポート ({period})"
WEEKLY_SHEET_DESCRIPTION = (
    "要件定義書(weekly_ai_intelligence_requirements.xlsx)のルールに基づき分類・採点。"
    "合計スコア降順。"
)


class ReportStoreError(Exception):
    """中間xlsx として扱えない要求（period 表記の誤り・行の不整合）。"""


@dataclass(frozen=True, slots=True)
class WrittenReport:
    """書き出した結果（監査ログ・ジョブの戻り値用）。

    Attributes:
        path: 正規名のパス（上書き済み）
        archived: 退避先。初回実行（退避対象が無い）なら `None`
        sheet: 書いたシート名
        rows: 書いたデータ行数
    """

    path: Path
    archived: Path | None
    sheet: str
    rows: int


class ReportStore:
    """中間xlsx の読み書き（`ArtifactStore` 経由。直接 `open()` しない）。

    `application.usecases.filter.HistoryReader` を満たすので、`FilterWorker` へ
    そのまま渡せる（`read_history()`）。
    """

    def __init__(self, store: ArtifactStore) -> None:
        self._store = store

    @property
    def store(self) -> ArtifactStore:
        return self._store

    # --- 書き出し ---------------------------------------------------------

    def write_weekly(
        self,
        *,
        period: str,
        articles: Sequence[Mapping[str, Any]],
        exclusions: Sequence[Mapping[str, Any]] = (),
        revision: int,
        run_id: str,
    ) -> WrittenReport:
        """週次ブックへ1週ぶんを書く（記事シートは作り直し／除外ログは append）。

        Args:
            period: `2026-W31`
            articles: 週次22列の行（**合計スコア降順で渡すこと**。§8.1）
            exclusions: 除外ログ6列の行
            revision: 実行時に固定していた config の revision（退避先の名前）
            run_id: ジョブ実行ID（同上）

        Returns:
            書き出した結果

        Raises:
            ReportStoreError: period が週次の表記でない／行が列定義と噛み合わない
        """
        parsed = self._parse(period, expect_weekly=True)
        path = self._store.weekly_report_path()
        workbook = self._load(path)

        sheet = self._reset_sheet(workbook, parsed.text, WEEKLY_ARTICLE_SHEET)
        sheet.cell(
            row=1,
            column=1,
            value=WEEKLY_SHEET_TITLE_FORMAT.format(period=parsed.text),
        )
        sheet.cell(row=2, column=1, value=WEEKLY_SHEET_DESCRIPTION)
        # 3行目は空行（§8.1）。openpyxl は書かなければ空のまま。
        self._write_table(sheet, WEEKLY_ARTICLE_SHEET, articles)

        log_sheet = self._ensure_sheet(
            workbook, EXCLUSION_LOG_SHEET_NAME, EXCLUSION_LOG_SHEET
        )
        self._append_rows(log_sheet, EXCLUSION_LOG_SHEET, exclusions)
        self._move_exclusion_log_last(workbook)

        archived = self._save(
            workbook, path, period=parsed.text, revision=revision, run_id=run_id
        )
        logger.info(
            "weekly report written"
            " (period=%s, articles=%d, exclusions=%d, archived=%s)",
            parsed.text,
            len(articles),
            len(exclusions),
            archived,
        )
        return WrittenReport(
            path=path, archived=archived, sheet=parsed.text, rows=len(articles)
        )

    def write_monthly(
        self,
        *,
        period: str,
        cases: Sequence[Mapping[str, Any]],
        revision: int,
        run_id: str,
    ) -> WrittenReport:
        """月次ブックへ1ヶ月ぶんを書く（シートは作り直し）。

        Args:
            period: `2026-07`
            cases: 月次8列の行（**`No` 昇順＝章グルーピング順で渡すこと**。§8.2）
            revision: 実行時に固定していた config の revision
            run_id: ジョブ実行ID

        Raises:
            ReportStoreError: period が月次の表記でない／`No` が昇順でない
        """
        parsed = self._parse(period, expect_weekly=False)
        _ensure_ascending_numbers(cases)

        path = self._store.monthly_cases_path()
        workbook = self._load(path)
        sheet = self._reset_sheet(workbook, parsed.text, MONTHLY_CASE_SHEET)
        self._write_table(sheet, MONTHLY_CASE_SHEET, cases)

        archived = self._save(
            workbook, path, period=parsed.text, revision=revision, run_id=run_id
        )
        logger.info(
            "monthly report written (period=%s, cases=%d, archived=%s)",
            parsed.text,
            len(cases),
            archived,
        )
        return WrittenReport(
            path=path, archived=archived, sheet=parsed.text, rows=len(cases)
        )

    def append_exclusions(
        self,
        *,
        period: str,
        exclusions: Sequence[Mapping[str, Any]],
        revision: int,
        run_id: str,
    ) -> WrittenReport:
        """除外ログだけを積む（月次実行など、週次シートを書かない場合）。"""
        parsed = self._parse(period)
        path = self._store.weekly_report_path()
        workbook = self._load(path)
        sheet = self._ensure_sheet(
            workbook, EXCLUSION_LOG_SHEET_NAME, EXCLUSION_LOG_SHEET
        )
        self._append_rows(sheet, EXCLUSION_LOG_SHEET, exclusions)
        self._move_exclusion_log_last(workbook)

        archived = self._save(
            workbook, path, period=parsed.text, revision=revision, run_id=run_id
        )
        return WrittenReport(
            path=path,
            archived=archived,
            sheet=EXCLUSION_LOG_SHEET_NAME,
            rows=len(exclusions),
        )

    # --- 読み戻し ---------------------------------------------------------

    def read_weekly(self, period: str) -> list[dict[str, Any]]:
        """週次シートのデータ行を読む（列名 → 値）。無ければ空。"""
        return self._read_sheet(
            self._store.weekly_report_path(), period, WEEKLY_ARTICLE_SHEET
        )

    def read_monthly(self, period: str) -> list[dict[str, Any]]:
        """月次シートのデータ行を読む（列名 → 値）。無ければ空。"""
        return self._read_sheet(
            self._store.monthly_cases_path(), period, MONTHLY_CASE_SHEET
        )

    def read_exclusions(self, period: str | None = None) -> list[dict[str, Any]]:
        """除外ログの行を読む（積まれた順）。

        Args:
            period: 指定するとその期間ぶんだけ返す（`GET /reports/{period}` の
                `summary.excluded`。T-27）。⚠️ **振り分けは `収集日` から行う**
                ので、`収集日` が空の行はどの期間にも入らない
                （`_exclusions_by_period()` の理由と同じ）。省略すると全行

        Returns:
            除外ログの行（列名 → 値）
        """
        if period is None:
            return self._read_sheet(
                self._store.weekly_report_path(),
                EXCLUSION_LOG_SHEET_NAME,
                EXCLUSION_LOG_SHEET,
            )
        return list(self._exclusions_by_period().get(period, ()))

    def read_history(self, periods: Sequence[str]) -> DedupHistory:
        """重複判定（T-18）の履歴を xlsx から組み立てる（T-18 申し送り①）。

        - 本編に載った記事: 週次シート／月次シート
        - 除外ログの記事: `除外ログ` シート（§11.1 が参照先に含めている）。
          **行は `収集日` しか持たない**ので、`weekly_period_of()` /
          `monthly_period_of()` で period を割り出して該当する期間ぶんだけ積む

        ⚠️ **順序がそのまま代表の優先順**（設計書 §6.3）。渡された period の順
        （呼び出し側は新しい順に渡す）で、各期間について「本編 → 除外ログ」の順に
        積む。本編に残っている記事を代表にしたいため。

        Args:
            periods: 参照範囲の period（新しい順）

        Returns:
            既出記事の履歴
        """
        excluded = self._exclusions_by_period()
        history = DedupHistory()
        for period in periods:
            try:
                parsed = parse_period(period)
            except PeriodError:
                logger.warning("履歴として読めない period を無視しました: %r", period)
                continue

            rows = (
                self.read_weekly(parsed.text)
                if parsed.is_weekly
                else self.read_monthly(parsed.text)
            )
            source_column = "ソース" if parsed.is_weekly else "出典"
            for row in rows:
                entry = _known_article(
                    row, parsed.text, KnownOrigin.PUBLISHED, source_column
                )
                if entry is not None:
                    history.append(entry)

            for row in excluded.get(parsed.text, ()):
                entry = _known_article(row, parsed.text, KnownOrigin.EXCLUDED, "ソース")
                if entry is not None:
                    history.append(entry)
        return history

    # --- 内部（ブックとシートの操作）--------------------------------------

    def _parse(self, period: str, *, expect_weekly: bool | None = None) -> Period:
        try:
            parsed = parse_period(period)
        except PeriodError as exc:
            raise ReportStoreError(str(exc)) from exc
        if expect_weekly is True and not parsed.is_weekly:
            raise ReportStoreError(f"週次 period が必要です: {period!r}")
        if expect_weekly is False and not parsed.is_monthly:
            raise ReportStoreError(f"月次 period が必要です: {period!r}")
        return parsed

    def _load(self, path: Path) -> Workbook:
        """既存のブックを読む。無ければ空のブック（既定シートは消す）。"""
        if self._store.exists(path):
            return load_workbook(BytesIO(self._store.read_bytes(path)))
        workbook = Workbook()
        workbook.remove(workbook.active)
        return workbook

    def _save(
        self, workbook: Workbook, path: Path, *, period: str, revision: int, run_id: str
    ) -> Path | None:
        """退避してから原子的に上書きする（設計判断B ／ T-02）。

        ⚠️ **退避が先。** 上書き後に退避すると、退避されるのは新しい内容になる。
        """
        archived = self._store.archive(
            path, period=period, revision=revision, run_id=run_id
        )
        buffer = BytesIO()
        workbook.save(buffer)
        self._store.write_bytes(path, buffer.getvalue())
        return archived

    @staticmethod
    def _reset_sheet(workbook: Workbook, name: str, layout: SheetLayout) -> Worksheet:
        """同名シートを作り直す（再実行で行が二重にならないように）。"""
        if name in workbook.sheetnames:
            workbook.remove(workbook[name])
        sheet = workbook.create_sheet(title=name)
        _write_header(sheet, layout)
        return sheet

    @staticmethod
    def _ensure_sheet(workbook: Workbook, name: str, layout: SheetLayout) -> Worksheet:
        """無ければヘッダつきで作る（除外ログ用。既存なら中身を残す）。"""
        if name in workbook.sheetnames:
            return workbook[name]
        sheet = workbook.create_sheet(title=name)
        _write_header(sheet, layout)
        return sheet

    @staticmethod
    def _move_exclusion_log_last(workbook: Workbook) -> None:
        """`除外ログ` を末尾へ（§8.1 のシート構成「各週シート ＋ 除外ログ」）。"""
        if EXCLUSION_LOG_SHEET_NAME not in workbook.sheetnames:
            return
        index = workbook.sheetnames.index(EXCLUSION_LOG_SHEET_NAME)
        offset = len(workbook.sheetnames) - 1 - index
        workbook.move_sheet(EXCLUSION_LOG_SHEET_NAME, offset=offset)

    @staticmethod
    def _write_table(
        sheet: Worksheet, layout: SheetLayout, rows: Sequence[Mapping[str, Any]]
    ) -> None:
        """データ行を先頭から書く（ヘッダは `_reset_sheet` が書いている）。"""
        for offset, values in enumerate(rows):
            _write_row(sheet, layout, layout.first_data_row + offset, values)

    @staticmethod
    def _append_rows(
        sheet: Worksheet, layout: SheetLayout, rows: Iterable[Mapping[str, Any]]
    ) -> None:
        """最終行の後ろへ積む（除外ログ）。"""
        next_row = max(sheet.max_row + 1, layout.first_data_row)
        for offset, values in enumerate(rows):
            _write_row(sheet, layout, next_row + offset, values)

    def _read_sheet(
        self, path: Path, name: str, layout: SheetLayout
    ) -> list[dict[str, Any]]:
        if not self._store.exists(path):
            return []
        workbook = load_workbook(BytesIO(self._store.read_bytes(path)), read_only=True)
        try:
            if name not in workbook.sheetnames:
                return []
            sheet = workbook[name]
            width = len(layout.columns)
            rows: list[dict[str, Any]] = []
            for cells in sheet.iter_rows(
                min_row=layout.first_data_row, values_only=True
            ):
                trimmed = list(cells[:width]) + [None] * max(0, width - len(cells))
                if all(cell is None or str(cell).strip() == "" for cell in trimmed):
                    continue  # 空行は読み飛ばす（末尾に空行が残ることがある）
                rows.append(parse_row(layout.columns, trimmed))
            return rows
        finally:
            workbook.close()

    def _exclusions_by_period(self) -> dict[str, list[dict[str, Any]]]:
        """除外ログを `収集日` から period ごとに振り分ける（週次・月次の両方）。

        ⚠️ **`収集日` が空の行は振り分けられないので履歴に入らない。** §12 の
        フォーマット不備で `収集日` ごと欠けた記事がそれに当たる（T-20 は空セルの
        まま記録する）。重複判定の取りこぼしになるが、**日付が無い行をどこかの週へ
        当てはめる方が誤り**なので、そのままにしてある。
        """
        buckets: dict[str, list[dict[str, Any]]] = {}
        for row in self.read_exclusions():
            collected = row.get("収集日")
            if not collected:
                continue
            try:
                day = date.fromisoformat(str(collected))
            except ValueError:
                logger.warning("除外ログの収集日を読めませんでした: %r", collected)
                continue
            buckets.setdefault(weekly_period_of(day), []).append(row)
            buckets.setdefault(monthly_period_of(day), []).append(row)
        return buckets


# --- 行の読み書き（列定義は T-07 だけが持つ）--------------------------------


def _write_header(sheet: Worksheet, layout: SheetLayout) -> None:
    for index, name in enumerate(header_row(layout.columns), start=1):
        sheet.cell(row=layout.header_row, column=index, value=name)


def _write_row(
    sheet: Worksheet, layout: SheetLayout, row: int, values: Mapping[str, Any]
) -> None:
    try:
        cells = format_row(layout.columns, values)
    except Exception as exc:  # T-07 の `ReportColumnError` を包んで場所を足す
        raise ReportStoreError(
            f"{sheet.title} の {row} 行目を組み立てられません: {exc}"
        ) from exc
    for index, cell in enumerate(cells, start=1):
        sheet.cell(row=row, column=index, value=cell)


def _ensure_ascending_numbers(cases: Sequence[Mapping[str, Any]]) -> None:
    """月次の `No` が昇順であることを確かめる（§8.2「昇順＝章グルーピング順」）。

    ⚠️ **ここで並べ替えない。** `No` の順序は章の束ね方そのもの（T-21）なので、
    ライタが黙って直すと「章がバラバラなのに番号だけ整っている」表になる。
    """
    numbers = [case.get("No") for case in cases]
    if any(number is None for number in numbers):
        raise ReportStoreError("月次の行に No がありません")
    if numbers != sorted(numbers):  # ty: ignore[invalid-argument-type]
        raise ReportStoreError(f"月次の No が昇順ではありません: {numbers}")


def _known_article(
    row: Mapping[str, Any], period: str, origin: KnownOrigin, source_column: str
) -> KnownArticle | None:
    """読み戻した行を履歴の1件へ（タイトルも URL も無い行は捨てる）。"""
    title = str(row.get("タイトル") or "").strip()
    url = str(row.get("URL") or "").strip()
    if not title and not url:
        return None
    return KnownArticle(
        title=title,
        url=url,
        source=str(row.get(source_column) or "").strip(),
        period=period,
        origin=origin,
    )


def known_columns(layout: SheetLayout) -> tuple[ReportColumn, ...]:
    """そのシートの列定義（テストと T-24 / T-25 が列順を確かめるのに使う）。"""
    return layout.columns


__all__ = [
    "WEEKLY_SHEET_DESCRIPTION",
    "WEEKLY_SHEET_TITLE_FORMAT",
    "ReportStore",
    "ReportStoreError",
    "WrittenReport",
    "known_columns",
]
