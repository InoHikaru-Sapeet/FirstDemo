"""ドライラン再フィルタ（T-29。設計判断C ／ 設計書 §3.3 ／ 仕様書 §7.3-5）。

HTTP を通さずにユースケースそのものを固定する。重点は4つ:

1. **採点をやり直さない。** AI クライアントを一切渡していない（渡す口も無い）。
   保存済みの採点結果へ、決定的な部分だけを当て直す
2. **編集可能パラメータの3分割が `EDITABLE_PATHS` を覆っている。** 分け忘れると
   「新しい編集項目を足したのにドライランが黙って無視する」状態になる
3. **試算できない変更は 422 で断る**（「効果ゼロ」として黙って通さない）
4. **正規の成果物を1バイトも触らない**（設計判断C）

⚠️ 「一度落ちた記事は戻ってこない」「月次は試算できない」という**限界も
テストで固定している**。限界が消えたら（＝分析結果を成果物に残すようにしたら）
これらのテストは落ちるべきで、そのとき初めて仕様として書き換える。
"""

import copy
import json
import os
from datetime import UTC, datetime, timedelta
from pathlib import Path
from typing import Any, cast

import pytest
from openpyxl import load_workbook

from adapter.config_repository import ConfigRepository, ConfigRevisionConflictError
from adapter.storage.artifact_store import ArtifactStore
from adapter.xlsx.report_writer import ReportStore
from application.usecases.dry_run import (
    DETERMINISTIC_PATHS,
    NOT_PREVIEWABLE_PATHS,
    RESCORE_REQUIRED_PATHS,
    RESULT_FILENAME,
    DryRunNotPreviewableError,
    DryRunPeriodError,
    DryRunUsecase,
    NoScoredDataError,
    new_dry_run_id,
    reapply_adoption_class,
    unclassified_editable_paths,
)
from application.usecases.update_config import EDITABLE_PATHS, apply_patch
from enterprise.entities.config import IntelligenceConfig
from enterprise.entities.config_validation import ConfigIssueCode
from enterprise.entities.report_columns import EXCLUSION_LOG_SHEET_NAME

PERIOD = "2026-W31"
INITIAL_CONFIG_PATH = (
    Path(__file__).parents[1] / "enterprise" / "data" / "config_initial.json"
)


@pytest.fixture(scope="session")
def initial_raw() -> dict[str, Any]:
    return json.loads(INITIAL_CONFIG_PATH.read_text(encoding="utf-8"))


@pytest.fixture
def config(initial_raw: dict[str, Any]) -> IntelligenceConfig:
    return IntelligenceConfig.model_validate(copy.deepcopy(initial_raw))


@pytest.fixture
def store(tmp_path: Path) -> ArtifactStore:
    return ArtifactStore(root=tmp_path / "artifacts", scratch_ttl_hours=24, tz=UTC)


@pytest.fixture
def reports(store: ArtifactStore) -> ReportStore:
    return ReportStore(store)


@pytest.fixture
def usecase(
    store: ArtifactStore, reports: ReportStore, config: IntelligenceConfig
) -> DryRunUsecase:
    """`config.json` を置いた状態のユースケース。

    ⚠️ **AI クライアントを渡す引数が無い**こと自体が「採点をやり直さない」の
    実体。ここに口が生えたら、この機能の前提が変わったということ。
    """
    store.write_text(store.config_path(), config.model_dump_json())
    repo = ConfigRepository(cast(Any, None), store)  # load() は DB を使わない
    return DryRunUsecase(repo=repo, store=store, reports=reports, tz=UTC)


def scored_row(
    *,
    title: str,
    total: int,
    reliability: int = 9,
    adoption_class: str = "参考情報",
    url: str | None = None,
) -> dict[str, Any]:
    """週次22列の1行。**6軸の和が `total` になるように作る**（T-19 の約束）。"""
    remainder = total - reliability
    axes = {
        "緊急性鮮度_点": min(10, max(0, remainder)),
        "アドバイザリー活用度_点": min(15, max(0, remainder - 10)),
        "AI業界市場インパクト_点": min(20, max(0, remainder - 25)),
        "実務活用可能性_点": min(20, max(0, remainder - 45)),
        "顧客関連度_点": min(25, max(0, remainder - 65)),
    }
    assert sum(axes.values()) + reliability == total, "テストデータの和が合わない"
    return {
        "収集日": "2026-07-28",
        "情報カテゴリ": "enterprise_ai_case",
        "タイトル": title,
        "一言要約": "AIエージェントを導入した。契約業務が自動化された。",
        "合計スコア": total,
        "信頼性_点": reliability,
        **axes,
        "レポート採用区分": adoption_class,
        "実務活用可能性": "すぐ活用",
        "顧客関連度": "直接関係",
        "信頼性": "高",
        "地域": ["日本"],
        "情報種別": "専門メディア報道",
        "業務領域": ["業務プロセス改革"],
        "業界": ["不動産"],
        "AIテーマ": ["AIエージェント"],
        "ソース": "ITmedia",
        "URL": url or f"https://example.com/{total}",
    }


def exclusion_row(title: str = "広告記事") -> dict[str, Any]:
    return {
        "収集日": "2026-07-28",
        "タイトル": title,
        "URL": "https://example.com/ad",
        "ソース": "PR TIMES",
        "除外区分": "完全除外",
        "除外理由": "アフィリエイト・広告色の強いツール紹介",
    }


def seed_weekly(
    reports: ReportStore,
    rows: list[dict[str, Any]],
    exclusions: list[dict[str, Any]] | None = None,
) -> None:
    reports.write_weekly(
        period=PERIOD,
        articles=rows,
        exclusions=exclusions or [],
        revision=1,
        run_id="job_seed",
    )


def read_detail(path: Path, sheet: str) -> list[list[Any]]:
    workbook = load_workbook(path, read_only=True)
    try:
        return [list(row) for row in workbook[sheet].iter_rows(values_only=True)]
    finally:
        workbook.close()


# =============================================================================
# 1. 編集可能パラメータの3分割（この分割が T-29 の設計そのもの）
# =============================================================================


def test_every_editable_path_is_classified() -> None:
    """⚠️ **§7.2 の編集可能パラメータを1つ残らず仕分けていること。**

    仕分け漏れがあると `_ensure_previewable()` はそれを「決定的でない」として
    断る——安全側ではあるが、**なぜ断られたのかが誰にも説明できない**
    （理由の文言も付かない）。編集項目を足したらここも足すこと。
    """
    assert unclassified_editable_paths() == frozenset()


def test_the_three_buckets_are_disjoint() -> None:
    """同じパスが2つの区分に入っていないこと（理由が2つ出る）。"""
    buckets = [DETERMINISTIC_PATHS, RESCORE_REQUIRED_PATHS, NOT_PREVIEWABLE_PATHS]
    union = DETERMINISTIC_PATHS | RESCORE_REQUIRED_PATHS | NOT_PREVIEWABLE_PATHS

    assert len(union) == sum(len(bucket) for bucket in buckets)
    assert union == EDITABLE_PATHS


def test_only_the_thresholds_are_replayable_without_the_ai() -> None:
    """⚠️ **受け付ける集合を明示で固定する。**

    ここが黙って広がると、「AI が付けた点が変わるはずの変更」を決定的に
    再適用してしまう＝**嘘の件数**を admin に見せることになる。広げるときは
    「その変更で保存済みの点数が変わらない」と言い切れる根拠を添えること。
    """
    assert DETERMINISTIC_PATHS == frozenset(
        {
            "tunable_thresholds.min_total_score_to_publish",
            "tunable_thresholds.min_reliability_score_to_publish",
            "tunable_thresholds.adoption_class_score_map.propose_next_meeting",
            "tunable_thresholds.adoption_class_score_map.reference_info",
            "tunable_thresholds.adoption_class_score_map.share_only",
        }
    )


def test_the_axis_weights_need_a_rescore() -> None:
    """軸の配点は**得点上限そのもの**。古い上限で付いた点は引き伸ばせない。"""
    assert "scoring_axes.*.weight" in RESCORE_REQUIRED_PATHS


def test_the_exclusion_rules_are_blocked_by_missing_facts() -> None:
    """⚠️ **理屈では再適用できるが、要る事実が残っていない。**

    除外判定（T-17）は事実（該当ルール番号・鮮度）と config だけで決まり、AI に
    `severity` も `enabled` も見せていない（T-19）ので、強度を変えても事実は
    動かない。それでも試算できないのは `ArticleFacts` がどの成果物にも
    書かれていないため。→ T-38。
    """
    assert {"exclusion_rules.*.enabled", "exclusion_rules.*.severity"} <= (
        NOT_PREVIEWABLE_PATHS
    )


# =============================================================================
# 2. 決定的な再適用（採否・採用区分）
# =============================================================================


def test_raising_the_threshold_drops_the_articles_below_it(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    seed_weekly(
        reports,
        [
            scored_row(title="残る", total=83),
            scored_row(title="ぎりぎり残る", total=62),
            scored_row(title="落ちる", total=61),
        ],
    )

    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 62}},
    )

    assert result.baseline.adopted == 3
    assert result.summary.adopted == 2
    assert result.summary.excluded == 1


def test_the_threshold_boundary_is_inclusive(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """しきい値ちょうどは**載せる**（T-21 の採否と同じ向き）。"""
    seed_weekly(reports, [scored_row(title="ちょうど", total=70)])

    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 70}},
    )

    assert result.summary.adopted == 1


def test_the_reliability_threshold_is_applied_too(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """採否の条件は2つ（合計・信頼性）。**片方でも下回れば除外**。"""
    seed_weekly(reports, [scored_row(title="合計は高いが信頼性が低い", total=83)])

    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_reliability_score_to_publish": 10}},
    )

    assert result.summary.adopted == 0
    assert result.summary.excluded == 1


def test_lowering_the_threshold_cannot_bring_back_excluded_articles(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """⚠️ **既知の限界を仕様として固定する。**

    除外ログ（6列）は点数を持たないので、しきい値を下げても「増える件数」は
    出せない。試算できるのは「今載っている記事のうち何件落ちるか」まで。
    ここが変わるのは分析結果を成果物に残したとき（→ T-38）。
    """
    seed_weekly(
        reports,
        [scored_row(title="採用済み", total=83)],
        [exclusion_row("低スコアで落ちていた記事")],
    )

    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 10}},
    )

    assert result.summary.adopted == result.baseline.adopted == 1
    assert result.summary.excluded == result.baseline.excluded == 1


def test_the_adoption_class_is_decided_again_from_the_candidate_map(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    """採用区分（§6.4）は候補 config の `adoption_class_score_map` で決め直す。"""
    seed_weekly(
        reports, [scored_row(title="83点", total=83, adoption_class="参考情報")]
    )

    result = usecase.execute(
        period=PERIOD,
        patch={
            "tunable_thresholds": {
                "adoption_class_score_map": {"propose_next_meeting": 80}
            }
        },
    )

    rows = read_detail(result.result_path, PERIOD)
    header = rows[3]
    data = rows[4]
    assert data[header.index("レポート採用区分")] == "次回定例で提案"


def test_a_downgraded_article_stays_downgraded(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """⚠️ `low_priority` による降格（§5.4）を勝手に取り消さないこと。

    降格したという事実は成果物に残らない（残るのは下げた後の値だけ）ので、
    現行 config で決め直した区分との差から復元する。
    """
    # 合計 83 → 現行の score_map では `参考情報`（70〜84）。保存値が `共有のみ` なら
    # 1段下がっている＝実行時に降格が当たっていた。
    seed_weekly(
        reports, [scored_row(title="降格済み", total=83, adoption_class="共有のみ")]
    )

    result = usecase.execute(
        period=PERIOD,
        patch={
            "tunable_thresholds": {
                "adoption_class_score_map": {"propose_next_meeting": 80}
            }
        },
    )

    rows = read_detail(result.result_path, PERIOD)
    header, data = rows[3], rows[4]
    # 候補 config では `次回定例で提案` になるところを、降格を引き継いで1段下げる。
    assert data[header.index("レポート採用区分")] == "参考情報"


@pytest.mark.parametrize(
    ("stored", "expected"),
    [("参考情報", "次回定例で提案"), ("共有のみ", "参考情報")],
)
def test_reapply_adoption_class_carries_the_downgrade(
    config: IntelligenceConfig, stored: str, expected: str
) -> None:
    candidate = apply_patch(
        config,
        {
            "tunable_thresholds": {
                "adoption_class_score_map": {"propose_next_meeting": 80}
            }
        },
    )

    assert reapply_adoption_class(stored, 83, config, candidate) == expected


def test_reapply_adoption_class_does_not_guess_when_it_cannot_explain_the_value(
    config: IntelligenceConfig,
) -> None:
    """⚠️ 説明の付かない保存値から降格を推測しない（1段では説明できない差）。"""
    candidate = apply_patch(
        config,
        {
            "tunable_thresholds": {
                "adoption_class_score_map": {"propose_next_meeting": 80}
            }
        },
    )

    # 合計 83 は現行では `参考情報`。保存値 `不採用` は1段下げでは説明できない。
    assert reapply_adoption_class("不採用", 83, config, candidate) == "次回定例で提案"


def test_the_total_is_recomputed_from_the_axis_scores(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """⚠️ 合計スコアの列を鵜呑みにしない（合計は6軸の和。T-19 の約束）。"""
    row = scored_row(title="合計列が壊れている", total=83)
    row["合計スコア"] = 5  # 軸点の和（83）と食い違う値

    seed_weekly(reports, [row])
    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 80}},
    )

    # 列の 5 ではなく和の 83 で判定される＝採用のまま。
    assert result.summary.adopted == 1
    rows = read_detail(result.result_path, PERIOD)
    header, data = rows[3], rows[4]
    assert data[header.index("合計スコア")] == 83


# =============================================================================
# 3. 試算できない変更は断る
# =============================================================================


@pytest.mark.parametrize(
    ("patch", "path", "code"),
    [
        (
            {"scoring_axes": [{"id": "reliability", "weight": 12}]},
            "scoring_axes.*.weight",
            ConfigIssueCode.RESCORE_REQUIRED,
        ),
        (
            {"tunable_thresholds": {"target_industries": ["不動産"]}},
            "tunable_thresholds.target_industries",
            ConfigIssueCode.RESCORE_REQUIRED,
        ),
        (
            {"exclusion_rules": [{"no": 11, "enabled": False}]},
            "exclusion_rules.*.enabled",
            ConfigIssueCode.NOT_PREVIEWABLE,
        ),
        (
            {"tunable_thresholds": {"dedup": {"lookback_weeks": 2}}},
            "tunable_thresholds.dedup.lookback_weeks",
            ConfigIssueCode.NOT_PREVIEWABLE,
        ),
        (
            {"tunable_thresholds": {"weekly": {"max_common_topics": 2}}},
            "tunable_thresholds.weekly.max_common_topics",
            ConfigIssueCode.NOT_PREVIEWABLE,
        ),
    ],
)
def test_changes_that_cannot_be_replayed_are_refused(
    usecase: DryRunUsecase,
    reports: ReportStore,
    patch: dict[str, Any],
    path: str,
    code: ConfigIssueCode,
) -> None:
    """⚠️ **「効果ゼロ」として黙って通さない。**

    通すと admin は「この変更は件数に影響しない」と読む。実際は「この機能では
    分からない」で、意味が逆になる。
    """
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    with pytest.raises(DryRunNotPreviewableError) as caught:
        usecase.execute(period=PERIOD, patch=patch)

    assert [(issue.path, issue.code) for issue in caught.value.issues] == [(path, code)]
    assert caught.value.issues[0].reason


def test_a_mixed_patch_is_refused_whole(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """試算できる変更と混ざっていても、部分的に試算して見せない。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    with pytest.raises(DryRunNotPreviewableError) as caught:
        usecase.execute(
            period=PERIOD,
            patch={
                "tunable_thresholds": {"min_total_score_to_publish": 62},
                "exclusion_rules": [{"no": 11, "enabled": False}],
            },
        )

    paths = [issue.path for issue in caught.value.issues]
    assert paths == ["exclusion_rules.*.enabled"]


def test_nothing_is_written_when_the_patch_is_refused(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    with pytest.raises(DryRunNotPreviewableError):
        usecase.execute(
            period=PERIOD, patch={"exclusion_rules": [{"no": 11, "enabled": False}]}
        )

    assert not store.scratch_root.exists()


def test_an_empty_patch_is_a_no_op_preview(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """何も変えない試算は現状の件数を返す（before/after の基準になる）。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    result = usecase.execute(period=PERIOD, patch={})

    assert result.changed_paths == ()
    assert result.summary == result.baseline


# =============================================================================
# 4. 入力の検証
# =============================================================================


def test_a_monthly_period_is_refused(usecase: DryRunUsecase) -> None:
    """⚠️ **限界を仕様として固定する。** 月次実行は採点済みの22列を残さない。"""
    with pytest.raises(DryRunPeriodError):
        usecase.execute(period="2026-07", patch={})


@pytest.mark.parametrize("period", ["", "2026", "2026-W00", "2026-W99", "../etc"])
def test_a_malformed_period_is_refused(usecase: DryRunUsecase, period: str) -> None:
    with pytest.raises(DryRunPeriodError):
        usecase.execute(period=period, patch={})


def test_a_period_without_scored_data_is_refused(usecase: DryRunUsecase) -> None:
    with pytest.raises(NoScoredDataError):
        usecase.execute(period=PERIOD, patch={})


def test_a_stale_base_revision_conflicts(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """`base_revision` を付けたら `PUT /config` と同じ楽観ロックが掛かる。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    with pytest.raises(ConfigRevisionConflictError):
        usecase.execute(period=PERIOD, patch={}, base_revision=99)


# =============================================================================
# 5. 隔離（設計判断C）
# =============================================================================


def test_the_output_lands_only_under_scratch(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    result = usecase.execute(period=PERIOD, patch={})

    assert store.scratch_root in result.result_path.parents
    assert result.result_path.name == RESULT_FILENAME
    assert result.directory.name == result.dry_run_id


def test_the_canonical_artifacts_are_untouched(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    """⚠️ **正規ファイルを1バイトも変えない**（設計判断C の中心）。

    上書きだけでなく**履歴退避も起こさない**（`_history/` が増えると、
    ドライランのたびに世代が1つ消えていく）。
    """
    seed_weekly(reports, [scored_row(title="記事", total=61)], [exclusion_row()])
    before = {
        path.name: path.read_bytes() for path in store.root.iterdir() if path.is_file()
    }
    history_before = sorted(p.name for p in store.history_root.rglob("*"))

    usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 80}},
    )

    after = {
        path.name: path.read_bytes() for path in store.root.iterdir() if path.is_file()
    }
    assert after == before
    assert sorted(p.name for p in store.history_root.rglob("*")) == history_before


def test_the_dry_run_output_is_not_a_servable_artifact(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    """`GET /files/{filename}`（全ロール可）の許可リストに載らないこと。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    result = usecase.execute(period=PERIOD, patch={})

    assert store.servable_path(result.result_path.name) is None
    assert not store.is_servable(result.result_path.name)


def test_expired_dry_runs_are_purged_on_the_next_one(
    usecase: DryRunUsecase, reports: ReportStore, store: ArtifactStore
) -> None:
    """TTL 経過分の自動削除（T-02）。掃除役のスケジューラは立てない。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])
    expired = store.dry_run_dir("dry_old")
    expired.mkdir(parents=True)
    (expired / RESULT_FILENAME).write_bytes(b"x")
    past = (datetime.now(tz=UTC) - timedelta(hours=25)).timestamp()
    os.utime(expired, (past, past))

    result = usecase.execute(period=PERIOD, patch={})

    assert not expired.exists()
    assert result.result_path.is_file()


def test_two_dry_runs_do_not_collide(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    first = usecase.execute(period=PERIOD, patch={})
    second = usecase.execute(period=PERIOD, patch={})

    assert first.dry_run_id != second.dry_run_id
    assert first.result_path.is_file() and second.result_path.is_file()


def test_the_dry_run_id_has_no_path_separator() -> None:
    """`scratch/dry-run/{id}/` のディレクトリ名になるので区切りを含めない。"""
    dry_run_id = new_dry_run_id(datetime(2026, 8, 17, 8, tzinfo=UTC))

    assert dry_run_id.startswith("dry_20260817-080000-")
    assert "/" not in dry_run_id and ".." not in dry_run_id


# =============================================================================
# 6. 明細（除外区分・除外理由つき）
# =============================================================================


def test_the_detail_lists_the_new_exclusions_with_a_reason(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """完了条件「明細（除外区分・除外理由つき）がダウンロードできる」の中身。"""
    seed_weekly(reports, [scored_row(title="落ちる記事", total=61)])

    result = usecase.execute(
        period=PERIOD,
        patch={"tunable_thresholds": {"min_total_score_to_publish": 62}},
    )

    rows = read_detail(result.result_path, EXCLUSION_LOG_SHEET_NAME)
    header, data = rows[0], rows[1]
    assert data[header.index("タイトル")] == "落ちる記事"
    assert data[header.index("除外区分")] == "低スコア/信頼性不足"
    # しきい値と実測値の両方が読めること（本番の除外ログと同じ文言）。
    assert data[header.index("除外理由")] == "合計スコア 61 < 62"


def test_the_detail_keeps_the_exclusions_that_were_already_there(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """元からの除外も明細に並ぶ（`GET /reports` の `excluded` と数え方を揃える）。"""
    seed_weekly(reports, [scored_row(title="残る", total=83)], [exclusion_row()])

    result = usecase.execute(period=PERIOD, patch={})

    rows = read_detail(result.result_path, EXCLUSION_LOG_SHEET_NAME)
    assert len(rows) == 2  # ヘッダ + 1件
    assert result.summary.excluded == 1


def test_the_detail_says_it_is_a_dry_run(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """⚠️ scratch から持ち出されても正規の成果物と取り違えないこと。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    result = usecase.execute(period=PERIOD, patch={})

    rows = read_detail(result.result_path, PERIOD)
    assert rows[0][0] is not None and "ドライラン" in rows[0][0]
    assert result.dry_run_id in rows[1][0]
    assert "採点はやり直していない" in rows[1][0]


def test_the_detail_uses_the_same_columns_as_the_real_sheet(
    usecase: DryRunUsecase, reports: ReportStore
) -> None:
    """列と並びは正規の週次シートと同じ（admin が同じ見方で読める）。"""
    seed_weekly(reports, [scored_row(title="記事", total=83)])

    result = usecase.execute(period=PERIOD, patch={})

    detail = read_detail(result.result_path, PERIOD)[3]
    canonical = read_detail(reports.store.weekly_report_path(), PERIOD)[3]
    assert detail == canonical
