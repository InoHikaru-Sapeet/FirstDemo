"""中間xlsx のライタ／リーダ（T-22 ／ 設計書 §2.2 ／ 仕様書 §8 ／ 設計判断B）。

重点:

- **週次は §8.1 の4行構成**（タイトル / 説明 / 空行 / ヘッダ）＋ 5行目からデータ
- **列順は T-07 の定義だけ**を見る（ここに写しを持たない）。multi は `;` 区切り
- **除外ログは append**、**記事シートは作り直し**（再実行で二重に並ばない）
- 月次は8列・`No` 昇順（並べ替えはしない＝章の順序を壊さない）
- **書き出した xlsx を読み戻すと元の行に戻る**（ラウンドトリップ）
- 設計判断B: 上書き前に `_history/{period}/{revision}_{run_id}/` へ退避
"""

from pathlib import Path
from typing import Any

import pytest
from openpyxl import load_workbook

from adapter.storage.artifact_store import ArtifactStore
from adapter.xlsx.report_writer import (
    WEEKLY_SHEET_DESCRIPTION,
    WEEKLY_SHEET_TITLE_FORMAT,
    ReportStore,
    ReportStoreError,
)
from enterprise.entities.report_columns import (
    EXCLUSION_LOG_COLUMNS,
    EXCLUSION_LOG_SHEET_NAME,
    MONTHLY_CASE_COLUMNS,
    MULTI_VALUE_SEPARATOR,
    PARAGRAPH_SEPARATOR,
    WEEKLY_ARTICLE_COLUMNS,
    WEEKLY_ARTICLE_SHEET,
    header_row,
)
from enterprise.services.dedup import KnownOrigin

WEEKLY_PERIOD = "2026-W31"
MONTHLY_PERIOD = "2026-07"
REVISION = 3
RUN_ID = "run-0001"


@pytest.fixture
def store(tmp_path: Path) -> ArtifactStore:
    return ArtifactStore(root=tmp_path)


@pytest.fixture
def reports(store: ArtifactStore) -> ReportStore:
    return ReportStore(store)


def weekly_row(
    *,
    title: str = "大手不動産がAIエージェントで契約業務を自動化",
    url: str = "https://example.com/news/1",
    source: str = "ITmedia",
    total: int = 83,
    collected_at: str = "2026-07-28",
) -> dict[str, Any]:
    """週次22列の1行（列名は T-07 の定義から引く）。"""
    values: dict[str, Any] = {
        "収集日": collected_at,
        "情報カテゴリ": "enterprise_ai_case",
        "タイトル": title,
        "一言要約": "AIエージェントを導入した。契約業務が自動化された。",
        "合計スコア": total,
        "緊急性鮮度_点": 8,
        "信頼性_点": 9,
        "アドバイザリー活用度_点": 12,
        "AI業界市場インパクト_点": 15,
        "実務活用可能性_点": 17,
        "顧客関連度_点": total - 61,
        "レポート採用区分": "参考情報",
        "実務活用可能性": "すぐ活用",
        "顧客関連度": "直接関係",
        "信頼性": "高",
        "地域": ["日本", "海外"],
        "情報種別": "専門メディア報道",
        "業務領域": ["業務プロセス改革"],
        "業界": ["不動産", "建設"],
        "AIテーマ": ["AIエージェント"],
        "ソース": source,
        "URL": url,
    }
    return values


def exclusion_row(
    *,
    title: str = "落とした記事",
    url: str = "https://example.com/dropped",
    category: str = "完全除外",
    collected_at: str = "2026-07-28",
) -> dict[str, Any]:
    return {
        "収集日": collected_at,
        "タイトル": title,
        "URL": url,
        "ソース": "個人ブログ",
        "除外区分": category,
        "除外理由": "真偽不明の噂・未確認情報",
    }


def monthly_row(no: int, *, chapter: str = "第1章 業務自動化") -> dict[str, Any]:
    return {
        "No": no,
        "トピック(章)": chapter,
        "企業・組織": ["A社", "B社"],
        "タイトル": f"事例{no}",
        "URL": f"https://example.com/case/{no}",
        "出典": "ITmedia（2026-07-27）",
        "掲載月": MONTHLY_PERIOD,
        "解説": ["事実。", "詳細。", "示唆。"],
    }


def sheet_values(store: ArtifactStore, path: Path, name: str) -> list[list[Any]]:
    workbook = load_workbook(path)
    try:
        return [list(row) for row in workbook[name].iter_rows(values_only=True)]
    finally:
        workbook.close()


# --- 週次：§8.1 の行構成 -----------------------------------------------------


def test_the_weekly_sheet_has_the_four_line_preamble(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """1行目タイトル / 2行目説明 / 3行目空行 / 4行目ヘッダ / 5行目以降データ。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row()],
        revision=REVISION,
        run_id=RUN_ID,
    )

    rows = sheet_values(store, store.weekly_report_path(), WEEKLY_PERIOD)

    assert rows[0][0] == WEEKLY_SHEET_TITLE_FORMAT.format(period=WEEKLY_PERIOD)
    assert rows[1][0] == WEEKLY_SHEET_DESCRIPTION
    assert all(cell is None for cell in rows[2])
    assert rows[3] == header_row(WEEKLY_ARTICLE_COLUMNS)
    assert len(rows) == WEEKLY_ARTICLE_SHEET.first_data_row


def test_the_weekly_columns_follow_the_definition(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """22列・順序厳守。列名はここに書かず T-07 から引く。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row()],
        revision=REVISION,
        run_id=RUN_ID,
    )

    rows = sheet_values(store, store.weekly_report_path(), WEEKLY_PERIOD)

    assert len(rows[3]) == len(WEEKLY_ARTICLE_COLUMNS) == 22
    assert rows[4][:5] == [
        "2026-07-28",
        "enterprise_ai_case",
        "大手不動産がAIエージェントで契約業務を自動化",
        "AIエージェントを導入した。契約業務が自動化された。",
        83,
    ]


def test_multi_values_use_the_configured_separator(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """週次の multi 列は `;` 区切り（§8.1）。区切り文字も T-07 が持つ。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row()],
        revision=REVISION,
        run_id=RUN_ID,
    )

    rows = sheet_values(store, store.weekly_report_path(), WEEKLY_PERIOD)
    region = rows[4][header_row(WEEKLY_ARTICLE_COLUMNS).index("地域")]

    assert region == MULTI_VALUE_SEPARATOR.join(["日本", "海外"])


def test_the_rows_are_written_in_the_order_given(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """整列は T-21 の責務。ライタは渡された順（＝合計スコア降順）をそのまま書く。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[
            weekly_row(title="高い", url="https://example.com/1", total=90),
            weekly_row(title="中", url="https://example.com/2", total=80),
            weekly_row(title="低い", url="https://example.com/3", total=70),
        ],
        revision=REVISION,
        run_id=RUN_ID,
    )

    written = reports.read_weekly(WEEKLY_PERIOD)

    assert [row["合計スコア"] for row in written] == [90, 80, 70]


def test_a_weekly_period_is_required(reports: ReportStore) -> None:
    with pytest.raises(ReportStoreError):
        reports.write_weekly(
            period=MONTHLY_PERIOD, articles=[], revision=REVISION, run_id=RUN_ID
        )


# --- 除外ログ：append --------------------------------------------------------


def test_the_exclusion_log_is_appended(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """⚠️ 除外ログは積み上げ（過去の記録を消さない）。"""
    reports.write_weekly(
        period="2026-W30",
        articles=[],
        exclusions=[exclusion_row(title="1件目")],
        revision=REVISION,
        run_id=RUN_ID,
    )
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[],
        exclusions=[exclusion_row(title="2件目"), exclusion_row(title="3件目")],
        revision=REVISION,
        run_id="run-0002",
    )

    logged = reports.read_exclusions()

    assert [row["タイトル"] for row in logged] == ["1件目", "2件目", "3件目"]
    rows = sheet_values(store, store.weekly_report_path(), EXCLUSION_LOG_SHEET_NAME)
    assert rows[0] == header_row(EXCLUSION_LOG_COLUMNS)
    assert len(rows[0]) == len(EXCLUSION_LOG_COLUMNS) == 6


def test_the_exclusion_log_is_the_last_sheet(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """§8.1 のシート構成「各週シート ＋ 除外ログ」。"""
    reports.write_weekly(
        period="2026-W30",
        articles=[],
        exclusions=[exclusion_row()],
        revision=REVISION,
        run_id=RUN_ID,
    )
    reports.write_weekly(
        period=WEEKLY_PERIOD, articles=[], revision=REVISION, run_id="run-0002"
    )

    workbook = load_workbook(store.weekly_report_path())
    try:
        assert workbook.sheetnames == [
            "2026-W30",
            WEEKLY_PERIOD,
            EXCLUSION_LOG_SHEET_NAME,
        ]
    finally:
        workbook.close()


def test_exclusions_can_be_appended_without_a_weekly_sheet(
    reports: ReportStore,
) -> None:
    """月次実行の除外も週次ブックの `除外ログ` へ積む（T-21 備考）。"""
    reports.append_exclusions(
        period=MONTHLY_PERIOD,
        exclusions=[exclusion_row(title="月次で落とした")],
        revision=REVISION,
        run_id=RUN_ID,
    )

    assert [row["タイトル"] for row in reports.read_exclusions()] == ["月次で落とした"]


# --- 月次：8列・No 昇順 ------------------------------------------------------


def test_the_monthly_sheet_has_eight_columns(
    reports: ReportStore, store: ArtifactStore
) -> None:
    reports.write_monthly(
        period=MONTHLY_PERIOD,
        cases=[monthly_row(1), monthly_row(2)],
        revision=REVISION,
        run_id=RUN_ID,
    )

    rows = sheet_values(store, store.monthly_cases_path(), MONTHLY_PERIOD)

    assert rows[0] == header_row(MONTHLY_CASE_COLUMNS)
    assert len(rows[0]) == len(MONTHLY_CASE_COLUMNS) == 8
    assert [row[0] for row in rows[1:]] == [1, 2]


def test_the_monthly_commentary_keeps_three_paragraphs(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """§8.2：解説は `\\n\\n` 区切りの3段落。連結は T-07 の列定義が持つ。"""
    reports.write_monthly(
        period=MONTHLY_PERIOD,
        cases=[monthly_row(1)],
        revision=REVISION,
        run_id=RUN_ID,
    )

    rows = sheet_values(store, store.monthly_cases_path(), MONTHLY_PERIOD)
    commentary = rows[1][header_row(MONTHLY_CASE_COLUMNS).index("解説")]

    assert commentary == PARAGRAPH_SEPARATOR.join(["事実。", "詳細。", "示唆。"])
    assert "・".join(["A社", "B社"]) in rows[1]


def test_an_out_of_order_no_is_rejected(reports: ReportStore) -> None:
    """⚠️ ライタが黙って並べ替えない（`No` の順序＝章の束ね方そのもの。§8.2）。"""
    with pytest.raises(ReportStoreError):
        reports.write_monthly(
            period=MONTHLY_PERIOD,
            cases=[monthly_row(2), monthly_row(1)],
            revision=REVISION,
            run_id=RUN_ID,
        )


def test_a_monthly_period_is_required(reports: ReportStore) -> None:
    with pytest.raises(ReportStoreError):
        reports.write_monthly(
            period=WEEKLY_PERIOD, cases=[], revision=REVISION, run_id=RUN_ID
        )


# --- ラウンドトリップ ---------------------------------------------------------


def test_a_weekly_row_survives_the_round_trip(reports: ReportStore) -> None:
    """write → read で元の dict に戻る（列順・型・multi の分解込み）。"""
    original = weekly_row()

    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[original],
        revision=REVISION,
        run_id=RUN_ID,
    )
    restored = reports.read_weekly(WEEKLY_PERIOD)

    assert restored == [original]
    assert list(restored[0]) == [column.name for column in WEEKLY_ARTICLE_COLUMNS]


def test_a_monthly_row_survives_the_round_trip(reports: ReportStore) -> None:
    original = monthly_row(1)

    reports.write_monthly(
        period=MONTHLY_PERIOD,
        cases=[original],
        revision=REVISION,
        run_id=RUN_ID,
    )
    restored = reports.read_monthly(MONTHLY_PERIOD)

    assert restored == [original]


def test_an_exclusion_row_survives_the_round_trip(reports: ReportStore) -> None:
    original = exclusion_row()

    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[],
        exclusions=[original],
        revision=REVISION,
        run_id=RUN_ID,
    )

    assert reports.read_exclusions() == [original]


def test_reading_a_missing_sheet_is_empty(reports: ReportStore) -> None:
    """まだ書いていない period を読んでも落ちない（初回実行）。"""
    assert reports.read_weekly(WEEKLY_PERIOD) == []
    assert reports.read_monthly(MONTHLY_PERIOD) == []
    assert reports.read_exclusions() == []


# --- 設計判断B：正規名は上書き＋履歴退避 -------------------------------------


def test_the_previous_version_is_archived_before_the_overwrite(
    reports: ReportStore, store: ArtifactStore
) -> None:
    """⚠️ 退避されるのは**上書き前**の内容（後に退避すると新しい方が残る）。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row(title="1回目")],
        revision=REVISION,
        run_id=RUN_ID,
    )
    written = reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row(title="2回目")],
        revision=4,
        run_id="run-0002",
    )

    assert written.archived is not None
    assert written.archived.parent.name == "4_run-0002"
    archived = load_workbook(written.archived)
    try:
        assert archived[WEEKLY_PERIOD].cell(row=5, column=3).value == "1回目"
    finally:
        archived.close()
    assert [row["タイトル"] for row in reports.read_weekly(WEEKLY_PERIOD)] == ["2回目"]


def test_the_first_write_has_nothing_to_archive(reports: ReportStore) -> None:
    written = reports.write_weekly(
        period=WEEKLY_PERIOD, articles=[], revision=REVISION, run_id=RUN_ID
    )

    assert written.archived is None


def test_rerunning_a_period_replaces_the_sheet(reports: ReportStore) -> None:
    """§14 冪等性：同じ period の再実行で記事が二重に並ばない。"""
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row(title="A"), weekly_row(title="B")],
        revision=REVISION,
        run_id=RUN_ID,
    )
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row(title="A")],
        revision=REVISION,
        run_id="run-0002",
    )

    assert [row["タイトル"] for row in reports.read_weekly(WEEKLY_PERIOD)] == ["A"]


def test_other_periods_are_kept_when_one_is_rewritten(reports: ReportStore) -> None:
    """固定名ファイルの中で、他の週のシートを巻き込まない（upsert）。"""
    reports.write_weekly(
        period="2026-W30",
        articles=[weekly_row(title="先週")],
        revision=REVISION,
        run_id=RUN_ID,
    )
    reports.write_weekly(
        period=WEEKLY_PERIOD,
        articles=[weekly_row(title="今週")],
        revision=REVISION,
        run_id="run-0002",
    )

    assert [row["タイトル"] for row in reports.read_weekly("2026-W30")] == ["先週"]
    assert [row["タイトル"] for row in reports.read_weekly(WEEKLY_PERIOD)] == ["今週"]


# --- 履歴の読み戻し（T-18 申し送り①）----------------------------------------


def test_the_history_covers_both_the_sheets_and_the_exclusion_log(
    reports: ReportStore,
) -> None:
    """§11.1 の参照先は「各週シート ＋ 除外ログ」。"""
    reports.write_weekly(
        period="2026-W30",
        articles=[weekly_row(title="先週の記事", url="https://example.com/kept")],
        exclusions=[
            exclusion_row(
                title="先週落とした記事",
                url="https://example.com/dropped",
                collected_at="2026-07-21",
            )
        ],
        revision=REVISION,
        run_id=RUN_ID,
    )

    history = reports.read_history(["2026-W30"])

    assert [entry.title for entry in history.entries] == [
        "先週の記事",
        "先週落とした記事",
    ]
    assert [entry.origin for entry in history.entries] == [
        KnownOrigin.PUBLISHED,
        KnownOrigin.EXCLUDED,
    ]


def test_the_history_keeps_the_requested_order(reports: ReportStore) -> None:
    """⚠️ 渡された順がそのまま代表の優先順（設計書 §6.3。T-18 申し送り②）。"""
    for period, title in (("2026-W29", "2週前"), ("2026-W30", "先週")):
        reports.write_weekly(
            period=period,
            articles=[weekly_row(title=title, url=f"https://example.com/{title}")],
            revision=REVISION,
            run_id=f"run-{period}",
        )

    history = reports.read_history(["2026-W30", "2026-W29"])

    assert [entry.title for entry in history.entries] == ["先週", "2週前"]


def test_the_history_ignores_periods_outside_the_scope(reports: ReportStore) -> None:
    reports.write_weekly(
        period="2026-W30",
        articles=[weekly_row(title="先週")],
        exclusions=[exclusion_row(collected_at="2026-07-21")],
        revision=REVISION,
        run_id=RUN_ID,
    )

    assert len(reports.read_history(["2026-W28"])) == 0


def test_the_monthly_history_uses_the_cases_sheet(reports: ReportStore) -> None:
    reports.write_monthly(
        period=MONTHLY_PERIOD,
        cases=[monthly_row(1)],
        revision=REVISION,
        run_id=RUN_ID,
    )

    history = reports.read_history([MONTHLY_PERIOD])

    assert [entry.title for entry in history.entries] == ["事例1"]
    # 月次に `ソース` 列は無いので `出典` を媒体名として持つ。
    assert history.entries[0].source == "ITmedia（2026-07-27）"


def test_an_undated_exclusion_row_is_not_placed_in_a_period(
    reports: ReportStore,
) -> None:
    """⚠️ `収集日` が空の行はどの週にも当てはめない（§12 不備で日付ごと欠けた記事）。"""
    reports.write_weekly(
        period="2026-W30",
        articles=[],
        exclusions=[
            {
                "収集日": None,
                "タイトル": "日付の無い記事",
                "URL": "https://example.com/x",
                "ソース": None,
                "除外区分": "フォーマット不備",
                "除外理由": "§12検証エラー",
            }
        ],
        revision=REVISION,
        run_id=RUN_ID,
    )

    assert len(reports.read_history(["2026-W30"])) == 0
    # 記録そのものは残る（本編にも除外ログにも無い記事を作らない。T-20）。
    assert len(reports.read_exclusions()) == 1


def test_a_bad_period_in_the_scope_is_skipped(
    reports: ReportStore, caplog: pytest.LogCaptureFixture
) -> None:
    with caplog.at_level("WARNING"):
        assert len(reports.read_history(["2026-13"])) == 0

    assert any("読めない period" in record.message for record in caplog.records)
