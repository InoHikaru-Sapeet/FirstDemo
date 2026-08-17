"""Config API（`GET /config` / `GET /config/history` / `PUT /config`。T-12・T-13）。

⚠️ **このテストの主眼は「config を admin 以外に一切露出しない」**（仕様書 §2
重要要件・§6.1、顧客指定）。単なる「編集できない」ではなく:

- 非 admin は **GET でも中身が返らない**（403 のみ・本文に config 由来の値なし）
- **エラーの形から config の存在有無・構造を推測できない**
  （`config.json` が有る場合と無い場合で、非 admin への応答が**完全に同一**）
- 非 admin の `PUT` は patch の中身に関係なく 403（項目名を含む 422 を返さない）

加えて T-13 の要:

- 保存時は必ず **T-05 のクロスフィールド検証**を通り、不正なら**拒否**する
  （**自動補正しない**＝設計判断A。拒否時は `config.json` が変わらない）
- 変更は **監査ログ**に残り、拒否されたときは残らない
"""

import asyncio
import copy
import json
import os
from collections.abc import AsyncIterator, Iterator
from datetime import UTC, datetime, timedelta
from io import BytesIO
from pathlib import Path
from typing import Any

import pytest
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from sqlalchemy.pool import NullPool

from adapter.config_repository import ConfigRepository
from adapter.database.base import Base
from adapter.database.models.audit_log import AuditEventType, AuditLog
from adapter.database.models.user import User
from adapter.http.fastapi.auth.dependencies import get_db_session
from adapter.http.fastapi.main import app
from adapter.storage.artifact_store import ArtifactStore
from adapter.xlsx.report_writer import ReportStore
from config import get_settings
from enterprise.entities.config import IntelligenceConfig
from enterprise.entities.principal import Role
from enterprise.services.password import hash_password
from enterprise.services.service_token import hash_service_token

PASSWORD = "correct horse battery staple"
ADMIN_EMAIL = "admin@sapeet.com"
VIEWER_EMAIL = "viewer@sapeet.com"
EDITOR_EMAIL = "editor@sapeet.com"

SERVICE_TOKEN = "service-token-for-tests"

# T-04 / T-05 / T-11 と同じ実データ（仕様書 §5.2 の確定 config）。
INITIAL_CONFIG_PATH = (
    Path(__file__).parents[1] / "enterprise" / "data" / "config_initial.json"
)


@pytest.fixture(scope="session")
def initial_raw() -> dict[str, Any]:
    return json.loads(INITIAL_CONFIG_PATH.read_text(encoding="utf-8"))


@pytest.fixture
def artifact_root(tmp_path: Path) -> Path:
    return tmp_path / "artifacts"


@pytest.fixture
def client(
    tmp_path: Path, artifact_root: Path, monkeypatch: pytest.MonkeyPatch
) -> Iterator[TestClient]:
    """テスト用 DB と成果物ルートに差し替えた API クライアント。"""
    monkeypatch.setenv("SESSION_COOKIE_SECURE", "false")
    monkeypatch.setenv("ARTIFACT_ROOT", str(artifact_root))
    monkeypatch.setenv("SERVICE_TOKEN_HASH", hash_service_token(SERVICE_TOKEN))
    get_settings.cache_clear()

    engine = create_async_engine(
        f"sqlite+aiosqlite:///{tmp_path / 'api.db'}", poolclass=NullPool
    )

    async def create_schema() -> None:
        async with engine.begin() as conn:
            await conn.run_sync(Base.metadata.create_all)

    asyncio.run(create_schema())

    maker = async_sessionmaker(engine, expire_on_commit=False)

    async def override_db() -> AsyncIterator[AsyncSession]:
        async with maker() as session:
            yield session

    app.dependency_overrides[get_db_session] = override_db
    with TestClient(app) as test_client:
        test_client._maker = maker  # type: ignore[attr-defined]
        yield test_client
    app.dependency_overrides.clear()
    get_settings.cache_clear()


def seed_user(client: TestClient, email: str, role: Role = Role.VIEWER) -> str:
    async def insert() -> str:
        now = datetime(2026, 8, 1, tzinfo=UTC)
        async with client._maker() as session:  # type: ignore[attr-defined]
            user = User(
                user_id=f"usr_{email}",
                email=email,
                display_name="テスト 花子",
                password_hash=hash_password(PASSWORD),
                role=role,
                is_active=True,
                created_at=now,
                updated_at=now,
                password_updated_at=now,
                failed_login_attempts=0,
                locked_until=None,
            )
            session.add(user)
            await session.commit()
            return user.user_id

    return asyncio.run(insert())


def login(client: TestClient, email: str) -> None:
    response = client.post("/auth/login", json={"email": email, "password": PASSWORD})
    assert response.status_code == 200, response.text


def login_as(client: TestClient, role: Role) -> str:
    email = {
        Role.ADMIN: ADMIN_EMAIL,
        Role.EDITOR: EDITOR_EMAIL,
        Role.VIEWER: VIEWER_EMAIL,
    }[role]
    user_id = seed_user(client, email, role=role)
    login(client, email)
    return user_id


def seed_config(client: TestClient, raw: dict[str, Any]) -> IntelligenceConfig:
    """`config.json` を revision=1 で作る（T-14 の初期投入相当）。"""
    settings = get_settings()

    async def create() -> IntelligenceConfig:
        async with client._maker() as session:  # type: ignore[attr-defined]
            repo = ConfigRepository(
                session,
                ArtifactStore.from_settings(settings),
                tz=settings.tzinfo,
            )
            return await repo.create_initial(IntelligenceConfig.model_validate(raw))

    return asyncio.run(create())


def config_file_text(artifact_root: Path) -> str:
    return (artifact_root / "config.json").read_text(encoding="utf-8")


def audit_rows(client: TestClient) -> list[AuditLog]:
    async def read() -> list[AuditLog]:
        async with client._maker() as session:  # type: ignore[attr-defined]
            return list(
                (await session.execute(select(AuditLog).order_by(AuditLog.at)))
                .scalars()
                .all()
            )

    return asyncio.run(read())


def config_update_rows(client: TestClient) -> list[AuditLog]:
    return [row for row in audit_rows(client) if row.event_type == "config_update"]


@pytest.fixture
def config(client: TestClient, initial_raw: dict[str, Any]) -> IntelligenceConfig:
    return seed_config(client, copy.deepcopy(initial_raw))


# 非 admin が config API を叩く経路（`system` はサービストークン経由）。
NON_ADMIN_CALLERS = ["viewer", "editor", "system", "anonymous"]


def authenticate_as(client: TestClient, caller: str) -> None:
    if caller == "anonymous":
        return
    if caller == "system":
        client.headers["Authorization"] = f"Bearer {SERVICE_TOKEN}"
        return
    login_as(client, Role(caller))


def expected_status(caller: str) -> int:
    """未認証は 401、認証済みで権限なしは 403（混ぜない。T-09・T-40）。"""
    return 401 if caller == "anonymous" else 403


# =============================================================================
# T-12: GET /config
# =============================================================================


def test_an_admin_reads_the_current_config(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)

    response = client.get("/config")

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["revision"] == config.meta.revision
    assert body["config"]["scoring_total"] == 100
    assert len(body["config"]["scoring_axes"]) == 6


def test_the_config_response_round_trips_through_the_model(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """返した中身がそのまま `IntelligenceConfig` として読める。

    T-31 の型生成が成り立つ前提。
    """
    login_as(client, Role.ADMIN)

    body = client.get("/config").json()

    assert IntelligenceConfig.model_validate(body["config"]) == config


def test_a_missing_config_is_404_for_an_admin(client: TestClient) -> None:
    """初期マイグレーション（T-14）前。admin だけがこの 404 に到達できる。"""
    login_as(client, Role.ADMIN)

    response = client.get("/config")

    assert response.status_code == 404
    assert response.json()["detail"]["error"] == "config_not_found"


# --- 認可（最重要要件）-------------------------------------------------------


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_only_an_admin_can_read_the_config(client: TestClient, caller: str) -> None:
    """⚠️ editor / viewer / system / 未認証は config を**読めない**。

    `system` が §6.2 で「内部のみ」なのは、パイプラインがファイルを直接読む
    経路のこと。**HTTP のレスポンス経路は持たない**（設計書 §3.1）。
    """
    seed_config(client, json.loads(INITIAL_CONFIG_PATH.read_text(encoding="utf-8")))
    authenticate_as(client, caller)

    for path in ("/config", "/config/history"):
        response = client.get(path)
        assert response.status_code == expected_status(caller), (
            f"{caller} {path}: {response.text}"
        )


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_a_denied_response_carries_no_config_content(
    client: TestClient, initial_raw: dict[str, Any], caller: str
) -> None:
    """⚠️ 拒否の本文に config 由来の値が1つも出ないこと。

    「編集不可」ではなく「中身を見せない」が要件なので、revision・項目名・
    しきい値・enum の日本語値のいずれも漏らさない。
    """
    seed_config(client, copy.deepcopy(initial_raw))
    authenticate_as(client, caller)

    leaks = [
        "revision",
        "scoring_axes",
        "min_total_score_to_publish",
        "exclusion_rules",
        "tunable_thresholds",
        initial_raw["tunable_thresholds"]["weekly"]["target_industries"][0],  # 不動産
        initial_raw["enums"]["adoption_class"][0],  # 次回定例で提案
        str(initial_raw["meta"]["revision"]),
    ]

    for path in ("/config", "/config/history"):
        text = client.get(path).text
        for leak in leaks:
            assert leak not in text, f"{caller} {path} が {leak!r} を漏らしている"


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_the_denial_is_identical_whether_or_not_the_config_exists(
    client: TestClient, initial_raw: dict[str, Any], caller: str
) -> None:
    """⚠️ **存在の秘匿**：応答の差から `config.json` の有無を推測させない。

    config 未作成なら 404、作成済みなら 403 —— のように状態依存にすると、
    非 admin が「config はもう作られている」と分かってしまう。認可を
    ハンドラの手前（依存関係）で終わらせている＝ファイルを読まないことが
    その担保で、このテストはそれを応答の同一性として固定する。
    """
    authenticate_as(client, caller)

    before = client.get("/config")
    before_history = client.get("/config/history")

    seed_config(client, copy.deepcopy(initial_raw))

    after = client.get("/config")
    after_history = client.get("/config/history")

    assert (before.status_code, before.text) == (after.status_code, after.text)
    assert (before_history.status_code, before_history.text) == (
        after_history.status_code,
        after_history.text,
    )
    assert before.status_code == expected_status(caller)


def test_a_demoted_admin_stops_seeing_the_config_without_re_login(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """降格は次のリクエストから効く（ロールを毎回 users 行から解決するため）。"""
    admin_id = login_as(client, Role.ADMIN)
    seed_user(client, "next@sapeet.com", role=Role.ADMIN)
    assert client.get("/config").status_code == 200

    assert (
        client.patch(f"/users/{admin_id}/role", json={"role": "viewer"}).status_code
        == 200
    )

    # Cookie はそのまま。再ログインしていない。
    response = client.get("/config")
    assert response.status_code == 403
    assert "scoring_axes" not in response.text


# --- GET /config/history -----------------------------------------------------


def test_history_returns_the_four_documented_fields(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)

    response = client.get("/config/history")

    assert response.status_code == 200
    items = response.json()["items"]
    assert len(items) == 1
    assert set(items[0]) == {"revision", "updated_at", "updated_by", "diff_summary"}
    assert items[0]["revision"] == 1


def test_history_never_carries_the_config_body(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ 履歴は config の複製にしない（設計書 §3.3 の items は4項目）。"""
    login_as(client, Role.ADMIN)
    client.put(
        "/config",
        json={
            "base_revision": 1,
            "patch": {"tunable_thresholds": {"min_total_score_to_publish": 55}},
        },
    )

    text = client.get("/config/history").text

    assert "config_snapshot" not in text
    assert "scoring_axes" not in text
    assert "exclusion_rules" not in text


def test_history_is_newest_first(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)
    client.put(
        "/config",
        json={
            "base_revision": 1,
            "patch": {"tunable_thresholds": {"min_total_score_to_publish": 55}},
        },
    )

    items = client.get("/config/history").json()["items"]

    assert [item["revision"] for item in items] == [2, 1]
    assert items[0]["diff_summary"] is not None
    assert items[0]["updated_by"] == f"admin:usr_{ADMIN_EMAIL}"


# --- OpenAPI（T-31 の型生成の入力）------------------------------------------


def test_openapi_documents_the_config_endpoints(client: TestClient) -> None:
    paths = client.get("/openapi.json").json()["paths"]

    assert set(paths["/config"]) >= {"get", "put"}
    assert "get" in paths["/config/history"]


def test_openapi_carries_the_config_response_schema(client: TestClient) -> None:
    """T-12 完了条件：レスポンススキーマが OpenAPI に出ること。"""
    schemas = client.get("/openapi.json").json()["components"]["schemas"]

    assert "ConfigResponse" in schemas
    assert "IntelligenceConfig" in schemas
    assert set(schemas["ConfigHistoryResponse"]["properties"]) == {"items"}


# =============================================================================
# T-13: PUT /config
# =============================================================================


def put_config(
    client: TestClient, patch: dict[str, Any], *, base_revision: int = 1
) -> Any:
    return client.put("/config", json={"base_revision": base_revision, "patch": patch})


def test_an_admin_updates_an_editable_parameter(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)

    response = put_config(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 55}}
    )

    assert response.status_code == 200, response.text
    body = response.json()
    assert body["revision"] == 2
    assert body["updated_by"] == f"admin:usr_{ADMIN_EMAIL}"
    assert body["updated_at"] is not None

    current = client.get("/config").json()
    assert current["revision"] == 2
    assert current["config"]["tunable_thresholds"]["min_total_score_to_publish"] == 55


@pytest.mark.parametrize(
    ("patch", "check"),
    [
        (
            {"exclusion_rules": [{"no": 11, "enabled": False}]},
            lambda cfg: cfg["exclusion_rules"][10]["enabled"] is False,
        ),
        (
            {"exclusion_rules": [{"no": 13, "severity": "full_exclude"}]},
            lambda cfg: cfg["exclusion_rules"][12]["severity"] == "full_exclude",
        ),
        (
            {
                "information_categories": [
                    {"id": "ai_implementation_ops", "priority": "high"}
                ]
            },
            lambda cfg: cfg["information_categories"][6]["priority"] == "high",
        ),
        (
            {"tunable_thresholds": {"dedup": {"title_similarity_threshold": 0.9}}},
            lambda cfg: (
                cfg["tunable_thresholds"]["dedup"]["title_similarity_threshold"] == 0.9
            ),
        ),
        (
            # 対象業界は複数可（T-46 Step 3）。業界の数だけ週刊 HTML が出る。
            {"tunable_thresholds": {"weekly": {"target_industries": ["製造", "金融"]}}},
            lambda cfg: (
                cfg["tunable_thresholds"]["weekly"]["target_industries"]
                == ["製造", "金融"]
            ),
        ),
        (
            {"tunable_thresholds": {"monthly": {"target_case_count": 12}}},
            lambda cfg: cfg["tunable_thresholds"]["monthly"]["target_case_count"] == 12,
        ),
        (
            {"tunable_thresholds": {"min_reliability_score_to_publish": 7}},
            lambda cfg: (
                cfg["tunable_thresholds"]["min_reliability_score_to_publish"] == 7
            ),
        ),
    ],
)
def test_the_editable_parameters_of_spec_7_2_go_through(
    client: TestClient,
    config: IntelligenceConfig,
    patch: dict[str, Any],
    check: Any,
) -> None:
    login_as(client, Role.ADMIN)

    response = put_config(client, patch)

    assert response.status_code == 200, response.text
    assert check(client.get("/config").json()["config"])


def test_scoring_weights_can_be_rebalanced(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """配点は「合計100を保ったまま」なら通る（設計判断A の順方向）。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client,
        {
            "scoring_axes": [
                {"id": "customer_relevance", "weight": 30},
                {"id": "reliability", "weight": 5},
            ]
        },
    )

    assert response.status_code == 200, response.text
    axes = {
        axis["id"]: axis["weight"]
        for axis in client.get("/config").json()["config"]["scoring_axes"]
    }
    assert axes["customer_relevance"] == 30
    assert axes["reliability"] == 5
    assert sum(axes.values()) == 100


# --- T-05 のクロスフィールド検証を必ず通す（最重要要件）----------------------


def test_a_weight_sum_other_than_100_is_rejected(
    client: TestClient, config: IntelligenceConfig, artifact_root: Path
) -> None:
    """⚠️ 設計判断A：合計100でなければ**保存を拒否**する（自動正規化しない）。"""
    login_as(client, Role.ADMIN)
    before = config_file_text(artifact_root)

    response = put_config(
        client, {"scoring_axes": [{"id": "customer_relevance", "weight": 30}]}
    )

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "validation_failed"
    assert [issue["code"] for issue in detail["issues"]] == ["weight_sum_mismatch"]
    assert detail["issues"][0]["path"] == "scoring_axes"

    # ⚠️ 補正して保存されていないこと。ファイルは1バイトも変わらない。
    assert config_file_text(artifact_root) == before
    assert client.get("/config").json()["revision"] == 1


def test_the_adoption_threshold_order_is_enforced(
    client: TestClient, config: IntelligenceConfig, artifact_root: Path
) -> None:
    """降順整合（§2.1.1-2）。設計書 §3.3 の例（60→62）がまさにこれに当たる。"""
    login_as(client, Role.ADMIN)
    before = config_file_text(artifact_root)

    response = put_config(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 62}}
    )

    assert response.status_code == 422, response.text
    codes = [issue["code"] for issue in response.json()["detail"]["issues"]]
    assert codes == ["adoption_threshold_order"]
    assert config_file_text(artifact_root) == before


def test_an_unknown_target_industry_is_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """参照整合（§2.1.1-3）。enum 外の業界は出力ファイル名に入り込む。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client,
        {"tunable_thresholds": {"weekly": {"target_industries": ["宇宙開発"]}}},
    )

    assert response.status_code == 422
    codes = [issue["code"] for issue in response.json()["detail"]["issues"]]
    assert codes == ["unknown_industry_reference"]


def test_a_duplicated_target_industry_is_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ 業界の数がそのまま生成物の数（T-46 Step 3）。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client,
        {"tunable_thresholds": {"weekly": {"target_industries": ["不動産", "不動産"]}}},
    )

    assert response.status_code == 422
    codes = [issue["code"] for issue in response.json()["detail"]["issues"]]
    assert codes == ["duplicate_industry_reference"]


def test_all_violations_come_back_at_once(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """一度の保存で複数の違反をまとめて直せること（T-05 の方針・T-34 の表示要件）。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client,
        {
            "scoring_axes": [{"id": "customer_relevance", "weight": 30}],
            "tunable_thresholds": {"weekly": {"target_industries": ["宇宙開発"]}},
        },
    )

    assert response.status_code == 422
    codes = {issue["code"] for issue in response.json()["detail"]["issues"]}
    assert codes == {"weight_sum_mismatch", "unknown_industry_reference"}


# --- 許可リスト（§7.2 の編集可能パラメータのみ）------------------------------


@pytest.mark.parametrize(
    "patch",
    [
        pytest.param({"scoring_total": 90}, id="scoring_total"),
        pytest.param({"schema_version": "2.0"}, id="schema_version"),
        pytest.param({"meta": {"revision": 99}}, id="meta"),
        pytest.param({"enums": {"industry": ["宇宙開発"]}}, id="enums"),
        pytest.param(
            {"required_tags": [{"id": "region", "label": "地域2"}]}, id="tags"
        ),
        pytest.param({"source_whitelist_hint": ["example.com"]}, id="whitelist"),
        pytest.param(
            {"scoring_axes": [{"id": "reliability", "label": "信頼度"}]},
            id="axis_label",
        ),
        pytest.param(
            {"scoring_axes": [{"id": "reliability", "criterion": "変更"}]},
            id="axis_criterion",
        ),
        pytest.param(
            {"information_categories": [{"id": "ai_agent_automation", "label": "X"}]},
            id="category_label",
        ),
        pytest.param(
            {"exclusion_rules": [{"no": 1, "name": "書き換え"}]}, id="rule_name"
        ),
    ],
)
def test_fixed_items_are_rejected(
    client: TestClient,
    config: IntelligenceConfig,
    artifact_root: Path,
    patch: dict[str, Any],
) -> None:
    """⚠️ 固定項目（ID系・`scoring_total`・`schema_version` 等）を含む patch は 422。"""
    login_as(client, Role.ADMIN)
    before = config_file_text(artifact_root)

    response = put_config(client, patch)

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "validation_failed"
    assert detail["issues"][0]["code"] == "field_not_editable"
    assert config_file_text(artifact_root) == before


def test_the_patch_cannot_dictate_the_revision(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ `meta` を patch から触れると楽観ロックが成立しない（T-11 と対）。"""
    login_as(client, Role.ADMIN)

    response = put_config(client, {"meta": {"revision": 99, "updated_by": "偽装"}})

    assert response.status_code == 422
    assert client.get("/config").json()["revision"] == 1


def test_an_unknown_key_is_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """タイポを黙って無視しない（設定したつもりで挙動が変わらない事故を防ぐ）。"""
    login_as(client, Role.ADMIN)

    response = put_config(client, {"tunable_thresholds": {"min_total_scoreee": 55}})

    assert response.status_code == 422
    assert response.json()["detail"]["issues"][0]["code"] == "unknown_field"


def test_an_unknown_array_element_is_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """存在しない軸 ID は 422（ID を足す経路にしない）。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client, {"scoring_axes": [{"id": "creativity", "weight": 10}]}
    )

    assert response.status_code == 422
    assert response.json()["detail"]["issues"][0]["code"] == "unknown_target"


def test_a_value_outside_the_model_range_is_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """モデル（T-04）の値域違反も 422。`issues` の形は他と同じ。"""
    login_as(client, Role.ADMIN)

    response = put_config(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 500}}
    )

    assert response.status_code == 422
    issue = response.json()["detail"]["issues"][0]
    assert issue["code"] == "invalid_value"
    assert issue["path"] == "tunable_thresholds.min_total_score_to_publish"


def test_unknown_request_fields_are_rejected(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)

    response = client.put(
        "/config", json={"base_revision": 1, "patch": {}, "force": True}
    )

    assert response.status_code == 422


# --- 楽観ロック（§6.3）------------------------------------------------------


def test_a_stale_base_revision_is_a_conflict(
    client: TestClient, config: IntelligenceConfig, artifact_root: Path
) -> None:
    login_as(client, Role.ADMIN)
    assert (
        put_config(
            client, {"tunable_thresholds": {"min_total_score_to_publish": 55}}
        ).status_code
        == 200
    )
    before = config_file_text(artifact_root)

    response = put_config(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 50}}
    )

    assert response.status_code == 409, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "revision_conflict"
    assert detail["current_revision"] == 2
    assert config_file_text(artifact_root) == before


def test_a_conflict_is_reported_before_field_level_issues(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ 古い base_revision なら 409 が先。

    admin が見ていない config に対して項目別 422 を返しても直しようがなく、
    正しい案内は「読み直して再保存」であるため。
    """
    login_as(client, Role.ADMIN)
    put_config(client, {"tunable_thresholds": {"min_total_score_to_publish": 55}})

    response = put_config(
        client, {"scoring_axes": [{"id": "customer_relevance", "weight": 30}]}
    )

    assert response.status_code == 409


# --- 認可（最重要要件）------------------------------------------------------


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_only_an_admin_can_update_the_config(
    client: TestClient, initial_raw: dict[str, Any], artifact_root: Path, caller: str
) -> None:
    seed_config(client, copy.deepcopy(initial_raw))
    before = config_file_text(artifact_root)
    authenticate_as(client, caller)

    response = put_config(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 55}}
    )

    assert response.status_code == expected_status(caller), response.text
    assert config_file_text(artifact_root) == before


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_a_denied_update_never_reveals_the_config_structure(
    client: TestClient, initial_raw: dict[str, Any], caller: str
) -> None:
    """⚠️ **エラーの形から構造を推測させない。**

    非 admin が固定項目・未知キーを含む patch を送っても、返るのは 403/401 の
    1種類だけ。「その項目は編集できません」「そのキーはありません」といった
    項目別 422 を返すと、config にどんなキーがあるかが分かってしまう。
    """
    seed_config(client, copy.deepcopy(initial_raw))
    authenticate_as(client, caller)

    responses = [
        put_config(client, {}),
        put_config(client, {"scoring_total": 90}),
        put_config(client, {"存在しないキー": 1}),
        put_config(client, {"tunable_thresholds": {"min_total_score_to_publish": 55}}),
        put_config(client, {"meta": {"revision": 99}}, base_revision=99),
    ]

    statuses = {response.status_code for response in responses}
    bodies = {response.text for response in responses}
    assert statuses == {expected_status(caller)}
    assert len(bodies) == 1, "patch の中身によって応答が変わっている"
    assert "scoring_total" not in bodies.pop()


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_a_denied_update_is_identical_whether_or_not_the_config_exists(
    client: TestClient, initial_raw: dict[str, Any], caller: str
) -> None:
    """⚠️ **存在の秘匿**：PUT でも 404/403 の差から存在を推測させない。"""
    authenticate_as(client, caller)
    patch = {"tunable_thresholds": {"min_total_score_to_publish": 55}}

    before = put_config(client, patch)
    seed_config(client, copy.deepcopy(initial_raw))
    after = put_config(client, patch)

    assert (before.status_code, before.text) == (after.status_code, after.text)


# --- 監査ログ（仕様書 §6.1）--------------------------------------------------


def test_a_successful_update_is_audited(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """who / when / diff(before→after) / revision を残す（仕様書 §6.1・§4.4）。"""
    login_as(client, Role.ADMIN)

    assert (
        put_config(
            client, {"tunable_thresholds": {"min_total_score_to_publish": 55}}
        ).status_code
        == 200
    )

    rows = config_update_rows(client)
    assert len(rows) == 1
    row = rows[0]
    assert row.actor == f"admin:usr_{ADMIN_EMAIL}"
    assert row.revision == 2
    assert row.target == "config.json"
    assert row.period is None
    assert row.at is not None
    assert row.diff == {
        "tunable_thresholds.min_total_score_to_publish": {"before": 60, "after": 55}
    }


def test_the_audit_diff_excludes_meta_noise(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """`meta.*` は毎回変わる。混ぜると変えた判断基準が埋もれる（T-11 と同方針）。"""
    login_as(client, Role.ADMIN)
    put_config(client, {"exclusion_rules": [{"no": 11, "enabled": False}]})

    row = config_update_rows(client)[0]

    assert row.diff is not None
    assert all(not path.startswith("meta.") for path in row.diff)


@pytest.mark.parametrize(
    ("patch", "base_revision"),
    [
        ({"scoring_axes": [{"id": "customer_relevance", "weight": 30}]}, 1),
        ({"scoring_total": 90}, 1),
        ({"tunable_thresholds": {"min_total_score_to_publish": 55}}, 99),
    ],
    ids=["cross_field_422", "not_editable_422", "conflict_409"],
)
def test_a_rejected_update_leaves_no_audit_log(
    client: TestClient,
    config: IntelligenceConfig,
    patch: dict[str, Any],
    base_revision: int,
) -> None:
    """⚠️ 「変えていないもの」を記録しない。監査ログは実際の変更と1:1で対応させる。"""
    login_as(client, Role.ADMIN)

    response = put_config(client, patch, base_revision=base_revision)

    assert response.status_code in (409, 422)
    assert config_update_rows(client) == []


def test_a_denied_update_leaves_no_audit_log(
    client: TestClient, config: IntelligenceConfig
) -> None:
    seed_user(client, VIEWER_EMAIL, role=Role.VIEWER)
    login(client, VIEWER_EMAIL)

    response = put_config(client, {"exclusion_rules": [{"no": 1, "enabled": False}]})

    assert response.status_code == 403
    assert audit_rows(client) == []


def test_the_audit_event_type_is_the_documented_one() -> None:
    assert AuditEventType.CONFIG_UPDATE.value == "config_update"


# =============================================================================
# T-29: POST /config/dry-run ／ GET /config/dry-run/{id}/result.xlsx
# =============================================================================
#
# ⚠️ **このセクションの主眼は2つ。**
#
# 1. **認可が config ファミリと同じであること**（設計書 §3.4）。dry-run は
#    未保存の config 値とその適用挙動を露出するので、`POST /run`（editor も 202）
#    ではなく config ファミリ（admin だけ）。明細ファイルも同じ扱い
# 2. **ドライランの成果物が正規の経路へ混ざらないこと**（設計判断C）。
#    `GET /reports/{period}` の件数・一覧も `GET /files/{filename}` も動かない
#
# 試算そのもの（何件になるか・限界）は `tests/application/test_dry_run.py` の担当。


DRY_RUN_PERIOD = "2026-W31"


def seed_weekly_sheet(client: TestClient, totals: list[int]) -> None:
    """その period の週次シート（採点済み22列）を作る。"""
    store = ArtifactStore.from_settings(get_settings())
    rows = [
        {
            "収集日": "2026-07-28",
            "情報カテゴリ": "enterprise_ai_case",
            "タイトル": f"記事{total}",
            "一言要約": "AIエージェントを導入した。契約業務が自動化された。",
            "合計スコア": total,
            "緊急性鮮度_点": 10,
            "信頼性_点": 9,
            "アドバイザリー活用度_点": 15,
            "AI業界市場インパクト_点": 20,
            "実務活用可能性_点": 20,
            "顧客関連度_点": total - 74,
            "レポート採用区分": "参考情報",
            "実務活用可能性": "すぐ活用",
            "顧客関連度": "直接関係",
            "信頼性": "高",
            "地域": ["日本"],
            "情報種別": "専門メディア報道",
            "業務領域": ["業務プロセス改革"],
            "業界": ["不動産"],
            "AIテーマ": ["AIエージェント"],
            "ソース": "ITmedia",
            "URL": f"https://example.com/{total}",
        }
        for total in totals
    ]
    ReportStore(store).write_weekly(
        period=DRY_RUN_PERIOD, articles=rows, revision=1, run_id="job_seed"
    )


def post_dry_run(
    client: TestClient,
    patch: dict[str, Any],
    period: str = DRY_RUN_PERIOD,
    **extra: Any,
) -> Any:
    body: dict[str, Any] = {"period": period, "candidate_config_patch": patch, **extra}
    return client.post("/config/dry-run", json=body)


def test_an_admin_previews_the_effect_of_an_unsaved_threshold(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83, 76, 75])

    response = post_dry_run(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 76}}
    )

    assert response.status_code == 202, response.text
    body = response.json()
    assert body["period"] == DRY_RUN_PERIOD
    assert body["base_revision"] == config.meta.revision
    assert body["baseline"] == {"adopted": 3, "excluded": 0}
    assert body["summary"] == {"adopted": 2, "excluded": 1}
    assert body["ttl_hours"] == get_settings().scratch_ttl_hours
    assert body["dry_run_id"].startswith("dry_")
    assert body["scratch_url"].endswith(f"/{body['dry_run_id']}/result.xlsx")


def test_the_detail_is_downloadable_from_the_url_the_response_gives(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """完了条件「明細（除外区分・除外理由つき）がダウンロードできる」。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83, 75])

    accepted = post_dry_run(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 80}}
    ).json()
    response = client.get(accepted["scratch_url"])

    assert response.status_code == 200, response.text
    assert response.headers["content-type"].startswith("application/vnd.openxmlformats")
    sheet = load_workbook(BytesIO(response.content))["除外ログ"]
    rows = [list(row) for row in sheet.iter_rows(values_only=True)]
    assert rows[1][4] == "低スコア/信頼性不足"
    assert rows[1][5] == "合計スコア 75 < 80"


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_only_an_admin_can_run_a_dry_run(
    client: TestClient, config: IntelligenceConfig, caller: str
) -> None:
    """⚠️ config ファミリ（設計書 §3.4）。**editor も 403**。"""
    seed_weekly_sheet(client, [83])
    authenticate_as(client, caller)

    response = post_dry_run(
        client, {"tunable_thresholds": {"min_total_score_to_publish": 76}}
    )

    assert response.status_code == expected_status(caller)
    assert "min_total_score_to_publish" not in response.text


@pytest.mark.parametrize("caller", NON_ADMIN_CALLERS)
def test_only_an_admin_can_download_a_dry_run_detail(
    client: TestClient, config: IntelligenceConfig, caller: str
) -> None:
    """⚠️ ファイル経由で §3.4 の判断を迂回させない。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])
    accepted = post_dry_run(client, {}).json()
    client.cookies.clear()
    authenticate_as(client, caller)

    response = client.get(accepted["scratch_url"])

    assert response.status_code == expected_status(caller)


def test_a_dry_run_does_not_change_what_the_reports_endpoint_lists(
    client: TestClient, config: IntelligenceConfig, artifact_root: Path
) -> None:
    """⚠️ **設計判断C の核心**：試算の出力が正規の一覧・件数へ混ざらないこと。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83, 75])
    store = ArtifactStore.from_settings(get_settings())
    store.write_text(store.weekly_html_path("不動産", DRY_RUN_PERIOD), "<html></html>")
    before = client.get(f"/reports/{DRY_RUN_PERIOD}").json()

    post_dry_run(client, {"tunable_thresholds": {"min_total_score_to_publish": 80}})

    assert client.get(f"/reports/{DRY_RUN_PERIOD}").json() == before
    assert before["summary"] == {"adopted": 2, "excluded": 0}


def test_a_dry_run_result_is_not_served_by_the_public_file_route(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """`GET /files/{filename}`（全ロール可）へ scratch を通さないこと。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])
    accepted = post_dry_run(client, {}).json()

    assert client.get("/files/result.xlsx").status_code == 404
    assert client.get(f"/files/{accepted['dry_run_id']}_result.xlsx").status_code == 404


def test_a_dry_run_leaves_the_canonical_workbook_untouched(
    client: TestClient, config: IntelligenceConfig, artifact_root: Path
) -> None:
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83, 75])
    canonical = artifact_root / "weekly_ai_intelligence_report.xlsx"
    before = canonical.read_bytes()
    config_before = config_file_text(artifact_root)

    post_dry_run(client, {"tunable_thresholds": {"min_total_score_to_publish": 80}})

    assert canonical.read_bytes() == before
    assert config_file_text(artifact_root) == config_before


def test_a_change_that_needs_a_rescore_is_refused_with_a_reason(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ 「効果ゼロ」と表示せず、**なぜ試算できないか**を返す。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])

    response = post_dry_run(
        client, {"scoring_axes": [{"id": "reliability", "weight": 12}]}
    )

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "not_previewable"
    assert detail["issues"][0]["path"] == "scoring_axes.*.weight"
    assert detail["issues"][0]["code"] == "rescore_required"


def test_a_change_whose_facts_were_never_stored_is_refused(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])

    response = post_dry_run(client, {"exclusion_rules": [{"no": 11, "enabled": False}]})

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "not_previewable"
    assert detail["issues"][0]["code"] == "not_previewable"


def test_a_patch_outside_the_allow_list_is_the_same_422_as_a_save(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """許可リスト違反は `PUT /config` と同じ封筒（フロントが同じ処理で扱える）。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])

    response = post_dry_run(client, {"scoring_total": 90})

    assert response.status_code == 422, response.text
    detail = response.json()["detail"]
    assert detail["error"] == "validation_failed"
    assert detail["issues"][0]["code"] == "field_not_editable"


def test_a_dry_run_on_a_period_without_scored_data_is_404(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)

    response = post_dry_run(client, {})

    assert response.status_code == 404, response.text
    assert response.json()["detail"]["error"] == "no_scored_data"


def test_a_monthly_period_is_refused_with_the_reason(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """月次実行は採点済みの22列を成果物に残さない（→ T-38）。"""
    login_as(client, Role.ADMIN)

    response = post_dry_run(client, {}, period="2026-07")

    assert response.status_code == 422, response.text
    assert response.json()["detail"]["error"] == "invalid_period"


def test_a_stale_base_revision_is_the_same_409_as_a_save(
    client: TestClient, config: IntelligenceConfig
) -> None:
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])

    response = post_dry_run(client, {}, base_revision=99)

    assert response.status_code == 409, response.text
    body = response.json()["detail"]
    assert body["error"] == "revision_conflict"
    assert body["current_revision"] == config.meta.revision


def test_an_unknown_dry_run_id_is_404(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """⚠️ 形が不正な ID も同じ 404（区別させない）。"""
    login_as(client, Role.ADMIN)

    assert client.get("/config/dry-run/dry_nope/result.xlsx").status_code == 404
    assert client.get("/config/dry-run/..%2F..%2Fetc/result.xlsx").status_code == 404


def test_an_expired_dry_run_is_gone(
    client: TestClient, config: IntelligenceConfig
) -> None:
    """TTL 経過分の自動削除（T-02 ／ 設計判断C）。"""
    login_as(client, Role.ADMIN)
    seed_weekly_sheet(client, [83])
    accepted = post_dry_run(client, {}).json()
    store = ArtifactStore.from_settings(get_settings())
    directory = store.dry_run_dir(accepted["dry_run_id"])
    past = (datetime.now(tz=UTC) - timedelta(hours=48)).timestamp()
    os.utime(directory, (past, past))

    assert client.get(accepted["scratch_url"]).status_code == 404
    assert not directory.exists()
